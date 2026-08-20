from __future__ import annotations

import json
import shutil
import unittest
from pathlib import Path

from build_delivery import build_delivery
from openpyxl import load_workbook


def write_jsonl(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("".join(json.dumps(row, ensure_ascii=False) + "\n" for row in rows), encoding="utf-8")


class BuildDeliveryTests(unittest.TestCase):
    def test_profile_batch_does_not_create_empty_comment_or_search_workbooks(self) -> None:
        out = Path(__file__).resolve().parent / ".xhs_batch_delivery_test_runtime"
        if out.exists():
            shutil.rmtree(out)
        out.mkdir()
        try:
            write_jsonl(out / "data" / "notes.jsonl", [{
                "note_id": "note-batch-1", "title": "合成列表卡片", "body": "",
                "author_id": "author-1", "author_name": "合成作者", "note_type": "image",
                "metrics": {"likes": 12}, "is_pinned": True,
                "canonical_url": "https://www.xiaohongshu.com/explore/note-batch-1",
                "field_scope": "visible_list_card_only", "detail_page_opened": False,
                "completion_state": "complete_visible_list_card", "collected_at": "2026-08-07T00:00:00Z",
            }])
            (out / "data" / "run_manifest.json").write_text(
                json.dumps({"state": "complete", "mode": "batch"}), encoding="utf-8"
            )
            summary = build_delivery(out)
            self.assertTrue((out / "01_笔记清单.xlsx").is_file())
            self.assertFalse((out / "02_评论明细.xlsx").exists())
            self.assertFalse((out / "03_搜索快照.xlsx").exists())
            self.assertEqual(summary["comment_workbook"], "")
            self.assertEqual(summary["search_workbook"], "")
        finally:
            if out.exists():
                shutil.rmtree(out)

    def test_single_note_includes_creator_snapshot_without_avatar(self) -> None:
        out = Path(__file__).resolve().parent / ".xhs_single_snapshot_test_runtime"
        if out.exists():
            shutil.rmtree(out)
        out.mkdir()
        try:
            write_jsonl(out / "data" / "notes.jsonl", [{
                "note_id": "note-single-1", "title": "合成单篇", "body": "仅用于测试",
                "author_id": "author-pseudonym", "author_name": "合成作者", "note_type": "image",
                "creator_snapshot": {
                    "nickname": "合成作者", "platform_account": "", "stable_creator_id": "profile-synthetic-1",
                    "profile_url": "https://www.xiaohongshu.com/user/profile/profile-synthetic-1",
                    "bio": "", "followers": None, "total_likes": None,
                    "snapshot_at": "2026-08-20T00:00:00Z",
                },
                "metrics": {"likes": 12}, "canonical_url": "https://www.xiaohongshu.com/explore/note-single-1",
                "completion_state": "complete_visible_note", "collected_at": "2026-08-20T00:00:00Z",
            }])
            (out / "data" / "run_manifest.json").write_text(
                json.dumps({"state": "complete", "mode": "all", "target_kind": "notes"}), encoding="utf-8"
            )
            build_delivery(out)
            workbook = load_workbook(out / "01_笔记清单.xlsx", read_only=True)
            try:
                self.assertIn("达人快照", workbook.sheetnames)
                self.assertEqual(workbook["达人快照"]["B2"].value, "合成作者")
                self.assertIsNone(workbook["达人快照"]["B7"].value)
                self.assertIn("未下载头像", workbook["达人快照"]["B12"].value)
            finally:
                workbook.close()
        finally:
            if out.exists():
                shutil.rmtree(out)

    def test_builds_three_workbooks_and_notes(self) -> None:
        out = Path(__file__).resolve().parent / ".xhs_delivery_test_runtime"
        if out.exists():
            shutil.rmtree(out)
        out.mkdir()
        try:
            write_jsonl(out / "data" / "notes.jsonl", [{
                "note_id": "note-1", "title": "合成笔记", "body": "仅用于测试", "author_id": "author-1",
                "author_name": "测试作者", "note_type": "image", "topics": ["#测试"], "mentions": [],
                "metrics": {"likes": 12, "collects": 3, "comments": 1, "shares": 0}, "is_pinned": False,
                "canonical_url": "https://www.xiaohongshu.com/explore/note-1", "collected_at": "2026-08-07T00:00:00Z",
                "completion_state": "complete_observed_note",
                "profile_id": "profile-1", "profile_rank": 1, "selection_reason": "pinned",
                "search_snapshot_id": "search-1", "search_rank": 1, "search_keyword": "测试",
            }])
            write_jsonl(out / "data" / "comments.jsonl", [{
                "comment_id": "comment-1", "comment_id_type": "platform", "note_id": "note-1", "level": 1,
                "author_id": "xhs_user_test", "content": "合成评论", "declared_reply_count": 0,
                "saved_reply_count": 0, "reply_expansion_status": "not_applicable",
            }])
            write_jsonl(out / "data" / "search_snapshots.jsonl", [{
                "search_snapshot_id": "search-1", "keyword": "测试", "tab": "全部", "filters": ["综合"],
                "captured_at": "2026-08-07T00:00:00Z", "state": "complete_first_n_visible_results",
                "results": [{"rank": 1, "note_id": "note-1", "title": "合成笔记", "author": "测试作者"}],
                "related_queries": ["相关词"],
            }])
            (out / "data" / "run_manifest.json").write_text(json.dumps({"state": "complete"}), encoding="utf-8")
            (out / "data" / "profile_selection.json").write_text(json.dumps({
                "profile_selection_id": "selection-1",
                "profile_id": "profile-1",
                "canonical_url": "https://www.xiaohongshu.com/user/profile/profile-1",
                "captured_at": "2026-08-07T00:00:00Z",
                "state": "complete_visible_pinned_plus_recent_n",
                "discovered_count": 6,
                "pinned_count": 1,
                "recent_requested": 5,
                "recent_selected": 5,
                "profile": {
                    "display_name": "合成账号",
                    "xiaohongshu_id": "synthetic-id",
                    "region_text": "测试地区",
                    "description": "仅用于测试",
                    "metrics": {"following": "1", "followers": "2", "likes_and_collects": "3"},
                },
                "selected": [{
                    "note_id": "note-1", "rank": 1, "is_pinned": True, "selection_reason": "pinned",
                    "title": "合成笔记", "author_name": "测试作者",
                    "canonical_url": "https://www.xiaohongshu.com/explore/note-1", "cover_url": "",
                }],
            }, ensure_ascii=False), encoding="utf-8")

            summary = build_delivery(out)
            self.assertEqual(summary["notes"], 1)
            self.assertTrue((out / "01_笔记清单.xlsx").is_file())
            self.assertTrue((out / "02_评论明细.xlsx").is_file())
            self.assertTrue((out / "03_搜索快照.xlsx").is_file())
            self.assertTrue((out / "05_采集说明.md").is_file())
            self.assertTrue(summary["profile_selection"])
            workbook = load_workbook(out / "01_笔记清单.xlsx", read_only=True, data_only=False)
            try:
                self.assertIn("账号信息", workbook.sheetnames)
                self.assertIn("主页选择", workbook.sheetnames)
                self.assertEqual(workbook["账号信息"]["B3"].value, "profile-1")
                self.assertEqual(workbook["主页选择"]["D2"].value, "pinned")
                headers = [cell.value for cell in workbook["笔记总览"][1]]
                self.assertIn("搜索快照ID", headers)
                self.assertIn("搜索位次", headers)
                self.assertIn("搜索关键词", headers)
            finally:
                workbook.close()
        finally:
            if out.exists():
                shutil.rmtree(out)


if __name__ == "__main__":
    unittest.main()

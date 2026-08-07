"""Build customer-readable workbooks from normalized Xiaohongshu JSONL files."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Iterable

from collector_core import atomic_write_json, utc_now


class DeliveryError(RuntimeError):
    pass


def _load_json(path: Path, default: Any) -> Any:
    if not path.is_file():
        return default
    return json.loads(path.read_text(encoding="utf-8-sig"))


def _load_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        return []
    rows: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        if line.strip():
            value = json.loads(line)
            if isinstance(value, dict):
                rows.append(value)
    return rows


def _flatten(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return "；".join(_flatten(item) for item in value if _flatten(item))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return str(value)


def _write_sheet(workbook: Any, title: str, headers: list[str], rows: Iterable[list[Any]]) -> None:
    sheet = workbook.create_sheet(title=title)
    sheet.append(headers)
    for row in rows:
        sheet.append([_flatten(value) for value in row])


def _style(workbook: Any) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor="0B6E75")
    font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = font
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in sheet.columns:
            letter = column[0].column_letter
            lengths = [len(str(cell.value or "")) for cell in column[:200]]
            sheet.column_dimensions[letter].width = min(max(max(lengths, default=8) + 2, 10), 48)
        if sheet.max_column == 2:
            sheet.column_dimensions["B"].width = min(max(sheet.column_dimensions["B"].width or 10, 64), 72)


def build_delivery(out_value: str | Path) -> dict[str, Any]:
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise DeliveryError("openpyxl is missing") from exc

    out = Path(out_value).expanduser().resolve()
    data = out / "data"
    notes = _load_jsonl(data / "notes.jsonl")
    comments = _load_jsonl(data / "comments.jsonl")
    assets = _load_jsonl(data / "assets.jsonl")
    searches = _load_jsonl(data / "search_snapshots.jsonl")
    run_manifest = _load_json(data / "run_manifest.json", {})
    profile_selection = _load_json(data / "profile_selection.json", {})
    if not notes and not comments and not searches:
        raise DeliveryError("No Xiaohongshu collection data was found")

    note_book = Workbook()
    note_book.remove(note_book.active)
    _write_sheet(note_book, "笔记总览", [
        "笔记ID", "主页ID", "主页位次", "选择原因", "标题", "正文", "作者ID", "作者", "类型", "发布时间原文", "地区原文",
        "点赞", "收藏", "评论", "分享", "是否置顶", "规范链接", "采集时间", "完成状态",
    ], [[
        row.get("note_id"), row.get("profile_id"), row.get("profile_rank"), row.get("selection_reason"),
        row.get("title"), row.get("body"), row.get("author_id"), row.get("author_name"), row.get("note_type"),
        row.get("published_at_text"), row.get("region_text"),
        (row.get("metrics") or {}).get("likes"), (row.get("metrics") or {}).get("collects"),
        (row.get("metrics") or {}).get("comments"), (row.get("metrics") or {}).get("shares"),
        row.get("is_pinned"), row.get("canonical_url"), row.get("collected_at"), row.get("completion_state"),
    ] for row in notes])
    if profile_selection:
        profile = profile_selection.get("profile") or {}
        metrics = profile.get("metrics") or {}
        _write_sheet(note_book, "账号信息", ["字段", "值"], [
            ["主页选择ID", profile_selection.get("profile_selection_id")],
            ["主页ID", profile_selection.get("profile_id")],
            ["账号名称", profile.get("display_name")],
            ["小红书号", profile.get("xiaohongshu_id")],
            ["地区原文", profile.get("region_text")],
            ["账号简介", profile.get("description")],
            ["关注原文", metrics.get("following")],
            ["粉丝原文", metrics.get("followers")],
            ["获赞与收藏原文", metrics.get("likes_and_collects")],
            ["规范主页链接", profile_selection.get("canonical_url")],
            ["采集时间", profile_selection.get("captured_at")],
            ["选择状态", profile_selection.get("state")],
            ["发现笔记数", profile_selection.get("discovered_count")],
            ["可见置顶数", profile_selection.get("pinned_count")],
            ["近期请求数", profile_selection.get("recent_requested")],
            ["近期选中数", profile_selection.get("recent_selected")],
            ["最终选中数", len(profile_selection.get("selected") or [])],
        ])
        _write_sheet(note_book, "主页选择", [
            "笔记ID", "主页位次", "是否置顶", "选择原因", "页面标题", "页面作者", "规范链接", "封面来源URL",
        ], [[
            row.get("note_id"), row.get("rank"), row.get("is_pinned"), row.get("selection_reason"),
            row.get("title"), row.get("author_name"), row.get("canonical_url"), row.get("cover_url"),
        ] for row in profile_selection.get("selected") or []])
    topic_rows = []
    for note in notes:
        for order, topic in enumerate(note.get("topics") or [], start=1):
            topic_rows.append([note.get("note_id"), "topic", order, topic])
        for order, mention in enumerate(note.get("mentions") or [], start=1):
            topic_rows.append([note.get("note_id"), "mention", order, mention])
    _write_sheet(note_book, "话题明细", ["笔记ID", "类型", "顺序", "原文"], topic_rows)
    _write_sheet(note_book, "素材索引", [
        "素材ID", "笔记ID", "类型", "顺序", "状态", "本地文件", "来源URL", "宽", "高", "字节数", "SHA256", "失败原因",
    ], [[
        row.get("asset_id"), row.get("note_id"), row.get("kind"), row.get("order"), row.get("status"),
        row.get("local_file"), row.get("source_url"), row.get("width"), row.get("height"), row.get("bytes"),
        row.get("sha256"), row.get("error_reason"),
    ] for row in assets])
    _write_sheet(note_book, "完整性", ["对象ID", "对象", "状态", "说明", "时间"], [[
        row.get("note_id"), "笔记", row.get("completion_state"), row.get("completion_note"), row.get("collected_at")
    ] for row in notes])
    _style(note_book)
    note_path = out / "01_笔记清单.xlsx"
    note_book.save(note_path)
    load_workbook(note_path, read_only=True).close()

    comment_book = Workbook()
    comment_book.remove(comment_book.active)
    _write_sheet(comment_book, "评论明细", [
        "评论ID", "ID类型", "笔记ID", "父评论ID", "根评论ID", "层级", "匿名作者ID", "作者显示名", "评论原文",
        "时间原文", "地区原文", "点赞原文", "声明回复数", "已保存回复数", "回复展开状态", "采集时间",
    ], [[
        row.get("comment_id"), row.get("comment_id_type"), row.get("note_id"), row.get("parent_comment_id"),
        row.get("root_comment_id"), row.get("level"), row.get("author_id"), row.get("author_display"), row.get("content"),
        row.get("time_text"), row.get("region_text"), row.get("like_count_text"), row.get("declared_reply_count"),
        row.get("saved_reply_count"), row.get("reply_expansion_status"), row.get("collected_at"),
    ] for row in comments])
    per_note: dict[str, dict[str, int]] = {}
    for row in comments:
        note = str(row.get("note_id") or "")
        item = per_note.setdefault(note, {"first": 0, "replies": 0})
        item["first" if int(row.get("level") or 1) == 1 else "replies"] += 1
    _write_sheet(comment_book, "采集状态", ["笔记ID", "保存一级评论", "保存回复", "完成状态"], [[
        note_id, value["first"], value["replies"], run_manifest.get("comment_states", {}).get(note_id, "unknown")
    ] for note_id, value in sorted(per_note.items())])
    _style(comment_book)
    comment_path = out / "02_评论明细.xlsx"
    comment_book.save(comment_path)
    load_workbook(comment_path, read_only=True).close()

    search_book = Workbook()
    search_book.remove(search_book.active)
    result_rows = []
    related_rows = []
    for snapshot in searches:
        for result in snapshot.get("results") or []:
            result_rows.append([
                snapshot.get("search_snapshot_id"), snapshot.get("keyword"), snapshot.get("tab"), snapshot.get("filters"),
                result.get("rank"), result.get("note_id") or result.get("result_note_id"), result.get("title") or result.get("result_title"),
                result.get("author") or result.get("result_author"), result.get("promoted_state"), snapshot.get("captured_at"), snapshot.get("state"),
            ])
        for order, query in enumerate(snapshot.get("related_queries") or [], start=1):
            related_rows.append([snapshot.get("search_snapshot_id"), order, query, snapshot.get("captured_at")])
    _write_sheet(search_book, "结果位次", [
        "搜索快照ID", "关键词", "标签页", "筛选条件", "位次", "笔记ID", "标题", "作者", "推广标记", "采集时间", "范围状态",
    ], result_rows)
    _write_sheet(search_book, "相关查询", ["搜索快照ID", "顺序", "相关查询", "采集时间"], related_rows)
    _style(search_book)
    search_path = out / "03_搜索快照.xlsx"
    search_book.save(search_path)
    load_workbook(search_path, read_only=True).close()

    (out / "04_笔记素材").mkdir(parents=True, exist_ok=True)
    description = f"""# BrandBAI 小红书采集说明

## 本次交付

- 笔记数：{len(notes)}
- 评论与回复记录：{len(comments)}
- 搜索快照数：{len(searches)}
- 运行状态：{run_manifest.get('state', 'unknown')}
- 主页选择状态：{profile_selection.get('state', 'not_applicable')}
- 主页选择范围：可见置顶 {profile_selection.get('pinned_count', 0)} 篇 + 最近非置顶 {profile_selection.get('recent_selected', 0)}/{profile_selection.get('recent_requested', 0)} 篇
- 构建时间：{utc_now()}

## 重要边界

1. 主页范围只对应采集时点页面可见置顶和最近非置顶笔记；置顶为额外项，不占最近 N 篇名额。
2. 搜索结果仅对应指定关键词、标签页、筛选和采集时点的页面可见顺序，不代表平台全量结果。
3. 评论完成只表示页面当前可返回内容收到终止信号；回复未逐楼展开时会保留部分完成状态。
4. 笔记与评论中的身份、购买、效果和体验主张未经本下载阶段核验。
5. 本包只做下载、结构化和质量说明，不自动生成用户语义、账号画像、内容机制、商品匹配或商业归因。
6. 登录资料、Cookie、请求头、验证码、页面临时令牌和本地任务缓存不进入交付包。
"""
    (out / "05_采集说明.md").write_text(description, encoding="utf-8")
    summary = {
        "built_at": utc_now(), "notes": len(notes), "comments": len(comments), "search_snapshots": len(searches),
        "run_state": run_manifest.get("state", "unknown"),
        "profile_selection": bool(profile_selection),
        "profile_selected_notes": len(profile_selection.get("selected") or []),
    }
    atomic_write_json(data / "delivery_manifest.json", summary)
    return summary


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--out", required=True)
    args = parser.parse_args()
    print(json.dumps(build_delivery(args.out), ensure_ascii=False, indent=2))

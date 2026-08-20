"""Build customer-readable TikTok workbooks from normalized JSONL files."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Iterable

from collector_core import atomic_write_json, utc_now


class DeliveryError(RuntimeError):
    pass


def _load_json(path: Path, default: Any) -> Any:
    return json.loads(path.read_text(encoding="utf-8-sig")) if path.is_file() else default


def _load_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        return []
    rows: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        if line.strip():
            value = json.loads(line)
            if isinstance(value, dict):
                rows.append(value)
    return rows


def _flat(value: Any) -> str | int | float:
    if value is None:
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    if isinstance(value, list):
        return "；".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return str(value)


def _sheet(book: Any, title: str, headers: list[str], rows: Iterable[list[Any]]) -> None:
    sheet = book.create_sheet(title=title)
    sheet.append(headers)
    for row in rows:
        sheet.append([_flat(value) for value in row])


def _style(book: Any) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    fill = PatternFill("solid", fgColor="111827")
    font = Font(color="FFFFFF", bold=True)
    for sheet in book.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = font
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in sheet.columns:
            letter = column[0].column_letter
            width = max((len(str(cell.value or "")) for cell in column[:200]), default=8) + 2
            sheet.column_dimensions[letter].width = min(max(width, 10), 48)


def build_delivery(out_value: str | Path) -> dict[str, Any]:
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise DeliveryError("openpyxl is missing") from exc
    out = Path(out_value).expanduser().resolve()
    data = out / "data"
    works = _load_jsonl(data / "works.jsonl")
    comments = _load_jsonl(data / "comments.jsonl")
    assets = _load_jsonl(data / "assets.jsonl")
    searches = _load_jsonl(data / "search_snapshots.jsonl")
    run = _load_json(data / "run_manifest.json", {})
    profile = _load_json(data / "profile_selection.json", {})
    selection = _load_json(data / "input_selection.json", {})
    if not works and not comments and not searches:
        raise DeliveryError("No TikTok collection data was found")

    book = Workbook()
    book.remove(book.active)
    _sheet(book, "作品总览", [
        "作品ID", "类型", "作者账号", "作者名称", "标题", "发布文案", "发布时间", "播放", "点赞",
        "评论", "收藏", "分享", "是否置顶", "选择原因", "规范链接", "采集时间", "完成状态",
    ], [[
        row.get("work_id"), row.get("work_type"), row.get("author_handle"), row.get("author_name"),
        row.get("title"), row.get("caption"), row.get("published_at"), (row.get("metrics") or {}).get("plays"),
        (row.get("metrics") or {}).get("likes"), (row.get("metrics") or {}).get("comments"),
        (row.get("metrics") or {}).get("collects"), (row.get("metrics") or {}).get("shares"),
        row.get("is_pinned"), row.get("selection_reason"), row.get("canonical_url"),
        row.get("collected_at"), row.get("completion_state"),
    ] for row in works])
    topic_rows = []
    for work in works:
        for order, topic in enumerate(work.get("hashtags") or [], start=1):
            topic_rows.append([work.get("work_id"), "hashtag", order, topic])
        for order, mention in enumerate(work.get("mentions") or [], start=1):
            topic_rows.append([work.get("work_id"), "mention", order, mention])
    _sheet(book, "话题明细", ["作品ID", "类型", "顺序", "原文"], topic_rows)
    _sheet(book, "素材索引", [
        "素材ID", "作品ID", "类型", "顺序", "状态", "本地文件", "公开来源指纹", "来源状态", "字节数", "SHA256", "失败原因",
    ], [[
        row.get("asset_id"), row.get("work_id"), row.get("kind"), row.get("order"), row.get("status"),
        row.get("local_file"), row.get("source_url"), row.get("source_url_state"), row.get("bytes"),
        row.get("sha256"), row.get("error_reason"),
    ] for row in assets])
    _sheet(book, "完整性", ["对象ID", "对象", "状态", "时间"], [
        [row.get("work_id"), "作品", row.get("completion_state"), row.get("collected_at")] for row in works
    ])
    creator_snapshot = works[0].get("creator_snapshot") if len(works) == 1 else None
    if isinstance(creator_snapshot, dict) and any(
        creator_snapshot.get(key) not in (None, "")
        for key in ("nickname", "platform_account", "stable_creator_id")
    ):
        _sheet(book, "达人快照", ["字段", "值"], [
            ["昵称", creator_snapshot.get("nickname")],
            ["TikTok账号", creator_snapshot.get("platform_account")],
            ["稳定达人ID", creator_snapshot.get("stable_creator_id")],
            ["主页链接", creator_snapshot.get("profile_url")],
            ["简介", creator_snapshot.get("bio")],
            ["粉丝数", creator_snapshot.get("followers")],
            ["累计获赞", creator_snapshot.get("total_likes")],
            ["快照时间", creator_snapshot.get("snapshot_at")],
            ["来源作品ID", works[0].get("work_id")],
            ["来源作品链接", works[0].get("canonical_url")],
            ["采集边界", "仅使用当前作品页已经展示或加载的公开作者信息；未自动进入达人主页；未下载头像。"],
        ])
    business_context = run.get("business_context") or {}
    if any(value not in (None, "") for value in business_context.values()):
        context_labels = {
            "business_preset": "业务预设", "market_scope": "目标市场", "source_surface": "来源界面",
            "source_locale": "页面语言／区域", "search_query_original": "原始搜索词",
            "search_language": "搜索语言", "observation_timezone": "观察时区",
            "authorization_mode": "授权模式", "downstream_use": "下游用途",
        }
        _sheet(book, "任务上下文", ["字段", "值"], [
            [context_labels.get(key, key), value] for key, value in business_context.items()
            if value not in (None, "")
        ])
    if profile:
        account = profile.get("profile") or {}
        _sheet(book, "账号信息", ["字段", "值"], [
            ["账号", profile.get("profile_handle")], ["显示名称", account.get("display_name")],
            ["规范主页链接", profile.get("canonical_url")], ["发现作品数", profile.get("discovered_count")],
            ["可见置顶数", profile.get("pinned_count")], ["近期请求数", profile.get("recent_requested")],
            ["近期选中数", profile.get("recent_selected")], ["选择状态", profile.get("state")],
            ["采集时间", profile.get("captured_at")],
        ])
        _sheet(book, "主页选择", ["作品ID", "位次", "类型", "置顶", "选择原因", "标题", "规范链接"], [[
            row.get("work_id"), row.get("rank"), row.get("work_type"), row.get("is_pinned"),
            row.get("selection_reason"), row.get("title"), row.get("url"),
        ] for row in profile.get("selected") or []])
    if selection:
        _sheet(book, "输入选择", [
            "选择顺序", "作品ID", "类型", "作者账号", "作者名称", "标题/发布文案", "是否置顶",
            "来源页面类型", "来源关键词", "来源排序", "规范链接", "选择原因",
        ], [[
            row.get("selection_rank"), row.get("work_id"), row.get("work_type"), row.get("author_handle"),
            row.get("author_name"), row.get("title"), row.get("is_pinned"), row.get("source_page_type"),
            row.get("source_keyword"), row.get("source_rank"), row.get("url"), row.get("selection_reason"),
        ] for row in selection.get("works") or []])
    _style(book)
    work_path = out / "01_作品清单.xlsx"
    out.mkdir(parents=True, exist_ok=True)
    book.save(work_path)
    load_workbook(work_path, read_only=True).close()

    comment_path = out / "02_评论明细.xlsx"
    if comments or run.get("comment_states"):
        comment_book = Workbook()
        comment_book.remove(comment_book.active)
        _sheet(comment_book, "评论明细", [
            "评论ID", "ID类型", "作品ID", "层级", "匿名作者ID", "作者显示名", "评论原文", "时间", "点赞",
            "声明回复数", "保存回复数", "回复展开状态", "采集时间",
        ], [[
            row.get("comment_id"), row.get("comment_id_type"), row.get("work_id"), row.get("level"),
            row.get("author_id"), row.get("author_display"), row.get("content"), row.get("create_time"),
            row.get("like_count"), row.get("declared_reply_count"), row.get("saved_reply_count"),
            row.get("reply_expansion_status"), row.get("collected_at"),
        ] for row in comments])
        _sheet(comment_book, "采集状态", ["作品ID", "保存一级评论", "完成状态"], [[
            work_id, len([row for row in comments if row.get("work_id") == work_id]), state
        ] for work_id, state in sorted((run.get("comment_states") or {}).items())])
        _style(comment_book)
        comment_book.save(comment_path)
        load_workbook(comment_path, read_only=True).close()
    else:
        comment_path.unlink(missing_ok=True)

    search_path = out / "03_搜索快照.xlsx"
    if searches:
        search_book = Workbook()
        search_book.remove(search_book.active)
        rows = []
        for snapshot in searches:
            for result in snapshot.get("results") or []:
                rows.append([snapshot.get("search_snapshot_id"), snapshot.get("keyword"), snapshot.get("tab"),
                             snapshot.get("filters"), result.get("rank"), result.get("work_id"), result.get("work_type"),
                             result.get("title"), result.get("url"), snapshot.get("captured_at"), snapshot.get("state")])
        _sheet(search_book, "结果位次", ["搜索快照ID", "关键词", "标签页", "筛选", "位次", "作品ID", "类型", "标题", "规范链接", "采集时间", "范围状态"], rows)
        _style(search_book)
        search_book.save(search_path)
        load_workbook(search_path, read_only=True).close()
    else:
        search_path.unlink(missing_ok=True)

    (out / "04_作品素材").mkdir(parents=True, exist_ok=True)
    asset_status_counts = run.get("asset_status_counts") or {
        status: len([row for row in assets if row.get("status") == status])
        for status in ("downloaded", "not_provided", "failed")
    }
    note = f"""# BrandBAI TikTok 采集说明

## 本次交付

- 作品数：{len(works)}
- 评论记录：{len(comments)}
- 素材记录：{len(assets)}
- 已下载素材：{asset_status_counts.get('downloaded', 0)}
- 公开页面未提供：{asset_status_counts.get('not_provided', 0)}
- 下载失败：{asset_status_counts.get('failed', 0)}
- 搜索快照数：{len(searches)}
- 运行状态：{run.get('state', 'unknown')}
- 主页选择状态：{profile.get('state', 'not_applicable')}
- 插件／文件选择状态：{selection.get('state', 'not_applicable')}
- 搜索选择状态：{run.get('search_selection_state', 'not_applicable')}
- 业务预设：{business_context.get('business_preset') or '未指定'}
- 目标市场：{business_context.get('market_scope') or '未指定'}
- 来源界面：{business_context.get('source_surface') or 'public_tiktok'}
- 页面语言／区域：{business_context.get('source_locale') or '未指定'}
- 原始搜索词：{business_context.get('search_query_original') or '不适用'}
- 搜索语言：{business_context.get('search_language') or '未指定'}
- 观察时区：{business_context.get('observation_timezone') or '未指定'}
- 授权模式：{business_context.get('authorization_mode') or 'public_visible'}
- 下游用途：{business_context.get('downstream_use') or '未指定'}
- 构建时间：{utc_now()}

## 重要边界

1. 主页范围只对应采集时点全部当前可见置顶和最近非置顶作品；置顶为额外项，不占最近 N 条名额。
2. 搜索结果只对应指定关键词、标签页、筛选和采集时点的页面可见顺序，不代表平台全量。
3. 评论只有收到分页终止或可信页面末端证据才标记完整；自动滚动本身不构成全量证明。
4. 回复未逐楼展开、达到正数上限或页面验证阻断时，保留部分完成状态。
5. 发布文案来自平台作品字段；Python 采集包不自动做语音转写、OCR 或云端翻译。本机双语证据可由支持该能力的 BrandBAI 下载助手生成。
6. 单作品包可包含“达人快照”；它不自动进入达人主页，不下载头像，未展示字段保持空白。
7. 本包只做下载、结构化和质量说明，不生成达人画像、用户语义、商品匹配、传播机制或销售归因。
8. Cookie、请求头、验证码、Chrome 资料夹和短期签名链接不进入交付包。
9. 作品没有独立原声地址时记录为“公开页面未提供”，不写成下载失败；若对应 MP4 已保存，可直接播放检查视频内嵌声音。
"""
    (out / "05_采集说明.md").write_text(note, encoding="utf-8")
    summary = {
        "built_at": utc_now(), "works": len(works), "comments": len(comments), "assets": len(assets),
        "search_snapshots": len(searches), "run_state": run.get("state", "unknown"),
        "explicit_selection_state": selection.get("state", "not_applicable"),
        "asset_status_counts": asset_status_counts,
        "business_context": business_context,
        "work_workbook": work_path.name,
        "comment_workbook": comment_path.name if comment_path.is_file() else "",
        "search_workbook": search_path.name if search_path.is_file() else "",
    }
    atomic_write_json(data / "delivery_manifest.json", summary)
    return summary


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", required=True)
    args = parser.parse_args()
    print(json.dumps(build_delivery(args.out), ensure_ascii=False, indent=2))

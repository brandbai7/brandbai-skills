import json
import tempfile
import unittest
from pathlib import Path

from build_delivery import DeliveryError, build_delivery


class BuildDeliveryTests(unittest.TestCase):
    def test_builds_conditional_workbooks(self):
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "delivery"
            data = out / "data"
            data.mkdir(parents=True)
            work = {"work_id": "7654321098765432101", "work_type": "photo", "author_handle": "demo",
                    "title": "synthetic", "caption": "synthetic #topic", "hashtags": ["topic"],
                    "metrics": {"plays": 10, "likes": 2, "comments": 1},
                    "canonical_url": "https://www.tiktok.com/@demo/photo/7654321098765432101",
                    "collected_at": "2026-01-01T00:00:00+00:00", "completion_state": "complete_visible_work"}
            comment = {"comment_id": "c1", "work_id": work["work_id"], "level": 1,
                       "author_id": "tiktok_user_x", "content": "synthetic comment", "collected_at": work["collected_at"]}
            snapshot = {"search_snapshot_id": "s1", "keyword": "test", "tab": "photo", "filters": ["relevance"],
                        "captured_at": work["collected_at"], "state": "complete_first_n_visible_results",
                        "results": [{"rank": 1, "work_id": work["work_id"], "work_type": "photo", "title": "synthetic",
                                     "url": work["canonical_url"]}]}
            (data / "works.jsonl").write_text(json.dumps(work) + "\n", encoding="utf-8")
            (data / "comments.jsonl").write_text(json.dumps(comment) + "\n", encoding="utf-8")
            (data / "search_snapshots.jsonl").write_text(json.dumps(snapshot) + "\n", encoding="utf-8")
            (data / "assets.jsonl").write_text("", encoding="utf-8")
            (data / "run_manifest.json").write_text(json.dumps({
                "state": "complete", "comment_states": {work["work_id"]: "complete_source_visible"},
                "business_context": {
                    "business_preset": "market-scan", "market_scope": "US",
                    "source_surface": "public_tiktok", "source_locale": "en-US",
                    "search_query_original": "test", "search_language": "en",
                    "observation_timezone": "America/New_York", "authorization_mode": "public_visible",
                    "downstream_use": "content-diagnosis",
                },
            }), encoding="utf-8")
            result = build_delivery(out)
            self.assertEqual(result["works"], 1)
            self.assertTrue((out / "01_作品清单.xlsx").is_file())
            self.assertTrue((out / "02_评论明细.xlsx").is_file())
            self.assertTrue((out / "03_搜索快照.xlsx").is_file())
            self.assertTrue((out / "05_采集说明.md").is_file())
            self.assertEqual(result["business_context"]["market_scope"], "US")
            from openpyxl import load_workbook
            book = load_workbook(out / "01_作品清单.xlsx", read_only=True)
            self.assertIn("任务上下文", book.sheetnames)
            book.close()
            self.assertIn("目标市场：US", (out / "05_采集说明.md").read_text(encoding="utf-8"))

    def test_refuses_empty_delivery(self):
        with tempfile.TemporaryDirectory() as tmp:
            with self.assertRaises(DeliveryError):
                build_delivery(tmp)


if __name__ == "__main__":
    unittest.main()

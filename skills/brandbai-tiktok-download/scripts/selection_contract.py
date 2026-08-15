"""Normalize explicit TikTok selections from JSON or BrandBAI Excel.

The contract contains metadata only. It never carries cookies, request headers,
browser profile data, access tokens, or transient signed media URLs. Work media
is enriched later through the normal visible signed-in TikTok page.
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse

from collector_core import canonical_handle, canonical_work_id, canonical_work_url, work_type_from_url


WORK_ROUTE_RE = re.compile(r"/@[^/]+/(?:video|photo)/(\d{10,})")
SUPPORTED_CONTRACTS = {"brandbai.tiktok.selection/v1", ""}


class SelectionContractError(RuntimeError):
    pass


def as_int(value: Any, default: int = 0) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return default


def optional_int(value: Any) -> int | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    try:
        return max(0, int(value))
    except (TypeError, ValueError):
        return None


def normalize_type(value: Any, source_url: str = "") -> str:
    text = str(value or "").strip().lower()
    if text in {"photo", "image", "image_post", "image-post", "图集", "照片", "图片"} or "/photo/" in source_url:
        return "photo"
    return "video"


def normalize_source_url(value: Any) -> str:
    source_url = str(value or "").strip()
    parsed = urlparse(source_url)
    if parsed.scheme != "https" or (parsed.hostname or "").lower() not in {"tiktok.com", "www.tiktok.com"}:
        raise SelectionContractError("Selection row must contain a valid https://www.tiktok.com work URL")
    if not WORK_ROUTE_RE.search(parsed.path):
        raise SelectionContractError("Selection row URL must contain /@handle/video/<id> or /@handle/photo/<id>")
    return source_url


def normalize_row(row: dict[str, Any], rank: int, defaults: dict[str, Any] | None = None) -> dict[str, Any]:
    defaults = defaults or {}
    source_url = normalize_source_url(
        row.get("url") or row.get("source_url") or row.get("canonical_url") or row.get("作品链接")
    )
    try:
        url_work_id = canonical_work_id(source_url)
    except ValueError as exc:
        raise SelectionContractError(f"Selection row {rank} has an invalid TikTok work URL") from exc
    work_id = str(row.get("work_id") or row.get("item_id") or row.get("作品ID") or url_work_id).strip()
    if work_id != url_work_id:
        raise SelectionContractError(f"Selection row {rank} work ID does not match its URL")
    work_type = normalize_type(row.get("work_type") or row.get("type") or row.get("类型"), source_url)
    if work_type != work_type_from_url(source_url):
        raise SelectionContractError(f"Selection row {rank} work type does not match its URL")
    metrics = row.get("metrics") if isinstance(row.get("metrics"), dict) else {}
    author_handle = str(row.get("author_handle") or row.get("作者账号") or "").strip().lstrip("@")
    if not author_handle:
        author_handle = canonical_handle(source_url)
    return {
        "work_id": work_id,
        "work_type": work_type,
        "author_handle": author_handle,
        "author_name": str(row.get("author_name") or row.get("author") or row.get("作者") or "").strip(),
        "title": str(row.get("title") or row.get("caption") or row.get("标题/发布文案") or "").strip(),
        "published_at": row.get("published_at") or row.get("发布时间") or None,
        "metrics": {
            "plays": optional_int(row.get("plays") if row.get("plays") is not None else row.get("播放数") if row.get("播放数") is not None else metrics.get("plays")),
            "likes": optional_int(row.get("likes") if row.get("likes") is not None else row.get("点赞数") if row.get("点赞数") is not None else metrics.get("likes")),
            "comments": optional_int(row.get("comments") if row.get("comments") is not None else row.get("评论数") if row.get("评论数") is not None else metrics.get("comments")),
            "collects": optional_int(row.get("collects") if row.get("collects") is not None else row.get("收藏数") if row.get("收藏数") is not None else metrics.get("collects")),
            "shares": optional_int(row.get("shares") if row.get("shares") is not None else row.get("分享数") if row.get("分享数") is not None else metrics.get("shares")),
        },
        "is_pinned": str(row.get("is_pinned") or row.get("是否置顶") or "").strip().lower() in {"1", "true", "yes", "是"},
        "url": canonical_work_url(source_url),
        "source_page_type": str(row.get("source_page_type") or row.get("来源页面类型") or defaults.get("page_type") or "selection").strip(),
        "source_keyword": str(row.get("source_keyword") or row.get("来源关键词") or defaults.get("keyword") or "").strip(),
        "source_rank": as_int(row.get("source_rank") or row.get("来源排序"), rank),
        "selection_reason": str(row.get("selection_reason") or defaults.get("selection_reason") or "插件作品清单").strip(),
        "selection_rank": rank,
    }


def deduplicate(rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows:
        work_id = str(row.get("work_id") or "")
        if not work_id or work_id in seen:
            continue
        seen.add(work_id)
        row["selection_rank"] = len(output) + 1
        output.append(row)
    return output


def load_json(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    try:
        payload: Any = json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError) as exc:
        raise SelectionContractError(f"Cannot read selection JSON: {path}") from exc
    metadata: dict[str, Any] = {}
    if isinstance(payload, dict):
        contract = str(payload.get("contract") or "").strip()
        if contract not in SUPPORTED_CONTRACTS:
            raise SelectionContractError(f"Unsupported selection contract: {contract}")
        source = payload.get("source") if isinstance(payload.get("source"), dict) else {}
        selection = payload.get("selection") if isinstance(payload.get("selection"), dict) else {}
        metadata = {
            "contract": contract or "brandbai.tiktok.selection/v1",
            "page_type": source.get("page_type") or "selection",
            "page_url": source.get("page_url") or "",
            "keyword": source.get("keyword") or "",
            "captured_at": source.get("captured_at") or "",
            "selection_mode": selection.get("mode") or "manual",
            "selection_reason": selection.get("label") or "插件作品清单",
        }
        rows = payload.get("works")
    else:
        rows = payload
    if not isinstance(rows, list):
        raise SelectionContractError("Selection JSON must contain a works list")
    normalized = deduplicate(
        normalize_row(row, index, metadata) for index, row in enumerate(rows, 1) if isinstance(row, dict)
    )
    if not normalized:
        raise SelectionContractError(f"No usable works found in selection JSON: {path}")
    return normalized, metadata


def load_xlsx(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SelectionContractError("openpyxl is required to read selection Excel files") from exc
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surface workbook parsing failures
        raise SelectionContractError(f"Cannot read selection Excel: {path}") from exc
    try:
        if "作品清单" not in workbook.sheetnames:
            raise SelectionContractError("Selection Excel must contain a 作品清单 sheet")
        sheet = workbook["作品清单"]
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(iterator, [])]
        for required in ("作品ID", "作品链接"):
            if required not in headers:
                raise SelectionContractError(f"作品清单 is missing the {required} column")
        raw_rows: list[dict[str, Any]] = []
        for values in iterator:
            row = {headers[index]: values[index] for index in range(min(len(headers), len(values))) if headers[index]}
            if str(row.get("作品ID") or "").strip():
                raw_rows.append(row)
        normalized = deduplicate(normalize_row(row, index) for index, row in enumerate(raw_rows, 1))
        if not normalized:
            raise SelectionContractError(f"No usable works found in selection Excel: {path}")
        page_types = {row["source_page_type"] for row in normalized if row.get("source_page_type")}
        keywords = {row["source_keyword"] for row in normalized if row.get("source_keyword")}
        return normalized, {
            "contract": "brandbai.tiktok.selection/v1",
            "page_type": next(iter(page_types)) if len(page_types) == 1 else "mixed",
            "page_url": "",
            "keyword": next(iter(keywords)) if len(keywords) == 1 else "",
            "captured_at": "",
            "selection_mode": "plugin_excel",
            "selection_reason": "插件作品清单",
        }
    finally:
        workbook.close()


def load_selection(path_value: str | Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    path = Path(path_value).expanduser().resolve()
    if not path.is_file():
        raise SelectionContractError(f"Selection file not found: {path}")
    if path.suffix.lower() == ".json":
        return load_json(path)
    if path.suffix.lower() == ".xlsx":
        return load_xlsx(path)
    raise SelectionContractError("--selection-file must be .json or .xlsx")

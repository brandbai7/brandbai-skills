"""Build the customer-readable BrandBAI Tmall collection workbooks."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Iterable

from collector_core import atomic_write_json, utc_now


class DeliveryError(RuntimeError):
    pass


def _load_json(path: Path, default: Any) -> Any:
    if not path.is_file():
        return default
    return json.loads(path.read_text(encoding="utf-8-sig"))


def _load_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not path.is_file():
        return rows
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        if not line.strip():
            continue
        value = json.loads(line)
        if isinstance(value, dict):
            rows.append(value)
    return rows


def _flatten(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return "；".join(_flatten(item) for item in value if _flatten(item))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return str(value)


def _style_workbook(workbook: Any) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill("solid", fgColor="173B57")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in sheet.columns:
            letter = column[0].column_letter
            lengths = [len(str(cell.value or "")) for cell in column[:200]]
            sheet.column_dimensions[letter].width = min(max(max(lengths, default=8) + 2, 10), 48)


def _write_sheet(workbook: Any, title: str, headers: list[str], rows: Iterable[list[Any]]) -> None:
    sheet = workbook.create_sheet(title=title)
    sheet.append(headers)
    for row in rows:
        sheet.append([_flatten(value) for value in row])


def build_delivery(out_value: str | Path) -> dict[str, Any]:
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise DeliveryError("openpyxl is missing; install requirements-browser.txt first") from exc

    out = Path(out_value).expanduser().resolve()
    product_root = out / "data" / "商品采集"
    review_root = out / "data" / "评价采集"
    question_root = out / "data" / "问答采集"
    products = [_load_json(path, {}) for path in sorted(product_root.glob("*/product.json"))] if product_root.is_dir() else []
    review_rows: list[dict[str, Any]] = []
    review_manifests: list[dict[str, Any]] = []
    if review_root.is_dir():
        for path in sorted(review_root.glob("*/reviews.jsonl")):
            review_rows.extend(_load_jsonl(path))
        review_manifests = [_load_json(path, {}) for path in sorted(review_root.glob("*/review_manifest.json"))]
    question_rows: list[dict[str, Any]] = []
    answer_rows: list[dict[str, Any]] = []
    question_manifests: list[dict[str, Any]] = []
    if question_root.is_dir():
        for path in sorted(question_root.glob("*/questions.jsonl")):
            question_rows.extend(_load_jsonl(path))
        for path in sorted(question_root.glob("*/answers.jsonl")):
            answer_rows.extend(_load_jsonl(path))
        question_manifests = [_load_json(path, {}) for path in sorted(question_root.glob("*/question_manifest.json"))]
    run_manifest = _load_json(out / "data" / "run_manifest.json", {})
    if not products and not review_rows and not review_manifests and not question_rows and not question_manifests:
        raise DeliveryError("No Tmall collection data was found in the delivery directory")

    product_book = Workbook()
    product_book.remove(product_book.active)
    _write_sheet(
        product_book,
        "商品总览",
        ["商品ID", "商品标题", "店铺", "商品链接", "选中SKU ID", "SKU映射状态", "当前选中规格", "页面参数适用范围", "规格参数冲突提示", "价格识别状态", "价格快照", "优惠权益", "销量快照", "库存快照", "有效详情内容图", "待复核低信息图", "模块读取状态", "详情加载状态", "详情加载步数", "详情位置已恢复", "视频识别状态", "页面视频播放器", "内容资料状态", "经营快照状态", "采集时间", "完成状态"],
        [
            [
                row.get("item_id"), row.get("title"), (row.get("shop") or {}).get("text"), row.get("canonical_url"),
                row.get("selected_sku_id"), row.get("sku_mapping_status"),
                [f"{item.get('name')}：{item.get('value')}" for item in (row.get("selected_sku_snapshot") or [])],
                row.get("parameter_scope_status"),
                f"{len(row.get('parameter_warnings') or [])} 项待人工确认" if row.get("parameter_warnings") else "未触发明确冲突",
                (row.get("snapshot") or {}).get("price_status"), (row.get("snapshot") or {}).get("price_texts"),
                (row.get("snapshot") or {}).get("benefit_texts"),
                (row.get("snapshot") or {}).get("sales_texts"), (row.get("snapshot") or {}).get("stock_texts"),
                row.get("effective_detail_image_count"),
                sum(1 for media in (row.get("media_records") or []) if media.get("kind") == "detail_image" and media.get("content_status") in {"separator_candidate", "low_information_candidate"}),
                "；".join(f"{name}:{state.get('status', 'unknown')}({state.get('count', 0)})" for name, state in (row.get("module_states") or {}).items()),
                row.get("detail_load_state"), row.get("detail_load_steps"), row.get("detail_scroll_restored"),
                (row.get("video_probe") or {}).get("status"), (row.get("video_probe") or {}).get("player_count"),
                row.get("material_status"), row.get("commerce_snapshot_status"),
                row.get("collected_at"), row.get("completion_state"),
            ] for row in products
        ],
    )
    parameter_rows = []
    sku_rows = []
    price_rows = []
    media_rows = []
    completion_rows = []
    warning_rows = []
    for product in products:
        item_id = product.get("item_id")
        for row in product.get("parameters") or []:
            parameter_rows.append([item_id, row.get("name"), row.get("value"), product.get("canonical_url"), product.get("collected_at")])
        for warning in product.get("parameter_warnings") or []:
            warning_rows.append([
                item_id, warning.get("selected_name"), warning.get("selected_value"),
                warning.get("parameter_name"), warning.get("parameter_value"),
                warning.get("reason"), "需人工确认",
            ])
        for group in product.get("sku_groups") or []:
            for order, value in enumerate(group.get("values") or [], start=1):
                sku_rows.append([
                    item_id, group.get("name"), order, value, value == group.get("selected_value"),
                    group.get("selected_value"), product.get("selected_sku_id"), product.get("collected_at"),
                ])
        for price in (product.get("snapshot") or {}).get("price_entries") or []:
            price_rows.append([
                item_id, price.get("role"), price.get("amount"), price.get("text"),
                price.get("context"), product.get("collected_at"),
            ])
        for row in product.get("media_records") or []:
            media_rows.append([
                item_id, row.get("asset_id"), row.get("kind"), row.get("order"), row.get("download_order"), row.get("content_status"), row.get("status"), row.get("file"),
                row.get("source_url"), row.get("source_url_query_redacted"), row.get("bytes"), row.get("content_type"),
            ])
        completion_rows.append([item_id, "商品资料", product.get("completion_state"), "", product.get("collected_at")])
    _write_sheet(product_book, "页面通用参数", ["商品ID", "参数名", "参数值", "来源链接", "采集时间"], parameter_rows)
    _write_sheet(product_book, "规格参数待确认", ["商品ID", "当前规格项", "当前规格值", "页面参数项", "页面参数值", "提示原因", "处理状态"], warning_rows)
    _write_sheet(product_book, "SKU快照", ["商品ID", "规格组", "顺序", "页面可见选项", "是否页面选中", "页面选中选项", "当时选中SKU ID", "采集时间"], sku_rows)
    _write_sheet(product_book, "价格与权益", ["商品ID", "价格角色", "金额", "页面原文", "局部上下文", "采集时间"], price_rows)
    _write_sheet(product_book, "素材索引", ["商品ID", "素材ID", "类型", "页面原序", "下载序", "内容判定", "状态", "文件", "公开来源URL", "查询参数已脱敏", "字节数", "内容类型"], media_rows)
    _write_sheet(product_book, "完整性", ["商品ID", "采集对象", "状态", "说明", "时间"], completion_rows)
    _style_workbook(product_book)
    product_path = out / "01_商品资料.xlsx"
    product_book.save(product_path)
    load_workbook(product_path, read_only=True).close()

    review_book = Workbook()
    review_book.remove(review_book.active)
    _write_sheet(
        review_book,
        "评价明细",
        ["评价ID", "ID类型", "商品ID", "内容角色", "评价者匿名ID", "页面遮罩名", "评价时间原文", "购买规格原文", "评价原文", "媒体数量", "采集时间"],
        [[
            row.get("review_id"), row.get("review_id_type"), row.get("item_id"), row.get("role"), row.get("author_id"),
            row.get("author_masked"), row.get("date_text"), row.get("purchased_sku_text"), row.get("content"),
            len(row.get("media") or []), row.get("collected_at"),
        ] for row in review_rows],
    )
    _write_sheet(
        review_book,
        "采集状态",
        ["商品ID", "完成状态", "已保存评价", "平台折叠提示数", "是否到达可见源末端", "样本上限", "是否达到上限", "隐私模式", "完成时间"],
        [[
            row.get("item_id"), row.get("state"), row.get("saved_reviews"), row.get("folded_count"), row.get("exhausted"),
            row.get("limit"), row.get("limit_reached"), row.get("privacy_mode"), row.get("finished_at"),
        ] for row in review_manifests],
    )
    _style_workbook(review_book)
    review_path = out / "02_评价明细.xlsx"
    review_book.save(review_path)
    load_workbook(review_path, read_only=True).close()

    question_book = Workbook()
    question_book.remove(question_book.active)
    _write_sheet(
        question_book,
        "问题清单",
        ["问题ID", "商品ID", "问题原文", "页面声明回答数", "商品链接", "采集时间"],
        [[
            row.get("question_id"), row.get("item_id"), row.get("content"), row.get("declared_answer_count"),
            row.get("canonical_url"), row.get("collected_at"),
        ] for row in question_rows],
    )
    _write_sheet(
        question_book,
        "回答明细",
        ["回答ID", "问题ID", "商品ID", "回答者匿名ID", "页面遮罩名", "购买或身份标签", "回答原文", "时间与规格原文", "采集时间"],
        [[
            row.get("answer_id"), row.get("question_id"), row.get("item_id"), row.get("author_id"),
            row.get("author_masked"), row.get("buyer_tag"), row.get("content"), row.get("meta_text"), row.get("collected_at"),
        ] for row in answer_rows],
    )
    _write_sheet(
        question_book,
        "采集状态",
        ["商品ID", "完成状态", "页面问题总数提示", "已保存问题", "已保存回答", "是否到达可见源末端", "样本上限", "是否达到上限", "完成时间"],
        [[
            row.get("item_id"), row.get("state"), row.get("total_hint"), row.get("saved_questions"),
            row.get("saved_answers"), row.get("exhausted"), row.get("limit"), row.get("limit_reached"), row.get("finished_at"),
        ] for row in question_manifests],
    )
    _style_workbook(question_book)
    question_path = out / "03_问大家.xlsx"
    question_book.save(question_path)
    load_workbook(question_path, read_only=True).close()

    if not products:
        product_path.unlink(missing_ok=True)
    if not review_rows and not review_manifests:
        review_path.unlink(missing_ok=True)
    if not question_rows and not question_manifests:
        question_path.unlink(missing_ok=True)

    description = f"""# BrandBAI 天猫商品资料采集说明

## 本次交付

- 商品数：{len(products)}
- 已保存评价及追评记录：{len(review_rows)}
- 已保存问大家问题：{len(question_rows)}
- 已保存问大家回答：{len(answer_rows)}
- 采集状态：{run_manifest.get('state', 'unknown')}
- 构建时间：{utc_now()}

## 文件说明

- `01_商品资料.xlsx`：选择商品资料模式时生成，包含商品总览、页面参数、SKU 映射、结构化价格与权益、素材索引和完整性状态。
- `02_评价明细.xlsx`：独立选择评价模式时生成，包含页面可见评价、追评与采集状态。
- `03_问大家.xlsx`：独立选择问大家模式时生成，包含问题、回答与采集状态。
- `03_商品素材/`：按本次选择下载的主图、详情图和可见视频。
- `data/`：用于断点续跑、复核和后续商品价值 Skill 接力的原始结构化数据。

## 重要边界

1. 价格、促销、销量、库存、排名和选中 SKU 都是采集时点快照，会随账号、地区、时间和活动变化。页面通用参数与当前选中规格分开保存，默认不宣称页面参数适用于当前 SKU。
2. 商品参数、主图、图文详情和视频按页面模块分别读取后合并；详情懒加载只在模块边界内有限前进、在推荐区前停止并恢复原位置，单模块失败不会清空其他模块成果。
3. 商品资料、用户评价和问大家是三个独立数据集；商品资料包不会因为评价或问答面板未打开而被阻塞。
4. “可见评价/问答采集完成”只表示滚动到页面当前可返回内容的末端；未打开完整面板或页面提示数量未覆盖时会保留部分完成状态。
5. 评价与问答是用户公开表达，不等于经过核验的商品功效、身份或购买事实。
6. 本包只做下载、结构化和质量说明，不自动生成商品价值结论、卖点、用户语义标签或商业归因。
7. 登录资料、Cookie、请求头、验证码、浏览器配置和本地任务缓存不进入交付包。
"""
    (out / "04_采集说明.md").write_text(description, encoding="utf-8")

    summary = {
        "built_at": utc_now(),
        "products": len(products),
        "reviews": len(review_rows),
        "questions": len(question_rows),
        "answers": len(answer_rows),
        "product_workbook": product_path.name if product_path.is_file() else "",
        "review_workbook": review_path.name if review_path.is_file() else "",
        "question_workbook": question_path.name if question_path.is_file() else "",
        "run_state": run_manifest.get("state", "unknown"),
    }
    atomic_write_json(out / "data" / "delivery_manifest.json", summary)
    return summary


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--out", required=True)
    args = parser.parse_args()
    print(json.dumps(build_delivery(args.out), ensure_ascii=False, indent=2))

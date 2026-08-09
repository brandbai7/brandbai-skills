from __future__ import annotations

import json
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from build_delivery import build_delivery
from package_delivery import package_directory


class BuildDeliveryTests(unittest.TestCase):
    def test_builds_workbooks_in_expected_plain_delivery_shape(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).resolve().parent) as temp:
            root = Path(temp) / "BrandBAI_天猫测试"
            item_id = "123456789"
            product_dir = root / "data" / "商品采集" / item_id
            review_dir = root / "data" / "评价采集" / item_id
            question_dir = root / "data" / "问答采集" / item_id
            product_dir.mkdir(parents=True)
            review_dir.mkdir(parents=True)
            question_dir.mkdir(parents=True)
            product = {
                "item_id": item_id,
                "title": "合成测试商品",
                "shop": {"text": "合成测试旗舰店", "href": "https://example.invalid/shop"},
                "canonical_url": f"https://detail.tmall.com/item.htm?id={item_id}",
                "selected_sku_id": "987654",
                "snapshot": {"price_texts": ["79.9"], "sales_texts": ["已售 100+"], "stock_texts": ["有货"]},
                "collected_at": "2026-08-07T00:00:00+00:00",
                "completion_state": "complete_observed_product",
                "parameters": [{"name": "品牌", "value": "合成品牌"}],
                "sku_groups": [{"name": "规格", "values": ["规格A", "规格B"], "selected_value": "规格A"}],
                "media_records": [{
                    "asset_id": f"tmall:{item_id}:main_image:001", "kind": "main_image", "order": 1,
                    "status": "downloaded", "file": "03_商品素材/合成测试商品/主图/001_main_image.webp",
                    "source_url": "https://img.alicdn.com/synthetic/a.webp", "source_url_query_redacted": True,
                    "bytes": 16, "content_type": "image/webp",
                }],
            }
            (product_dir / "product.json").write_text(json.dumps(product, ensure_ascii=False), encoding="utf-8")
            (product_dir / "asset_manifest.json").write_text(json.dumps(product["media_records"], ensure_ascii=False), encoding="utf-8")
            review = {
                "review_id": "derived:synthetic", "review_id_type": "derived", "item_id": item_id,
                "role": "review", "author_id": "reviewer_123", "author_masked": "", "date_text": "2026-08-01",
                "purchased_sku_text": "已购：规格A", "content": "这是一条纯合成测试评价。", "media": [],
                "collected_at": "2026-08-07T00:00:00+00:00",
            }
            (review_dir / "reviews.jsonl").write_text(json.dumps(review, ensure_ascii=False) + "\n", encoding="utf-8")
            review_manifest = {
                "item_id": item_id, "state": "partial_platform_folded", "saved_reviews": 1,
                "folded_count": 10, "exhausted": True, "limit": 0, "limit_reached": False,
                "privacy_mode": "pseudonymized", "finished_at": "2026-08-07T00:00:00+00:00",
            }
            (review_dir / "review_manifest.json").write_text(json.dumps(review_manifest, ensure_ascii=False), encoding="utf-8")
            question = {
                "question_id": "question:synthetic", "item_id": item_id,
                "content": "这是一条纯合成测试问题。", "declared_answer_count": 1,
                "canonical_url": f"https://detail.tmall.com/item.htm?id={item_id}",
                "collected_at": "2026-08-07T00:00:00+00:00",
            }
            answer = {
                "answer_id": "answer:synthetic", "question_id": "question:synthetic", "item_id": item_id,
                "author_id": "qa_reviewer_123", "author_masked": "", "buyer_tag": "已购",
                "content": "这是一条纯合成测试回答。", "meta_text": "规格A",
                "collected_at": "2026-08-07T00:00:00+00:00",
            }
            (question_dir / "questions.jsonl").write_text(json.dumps(question, ensure_ascii=False) + "\n", encoding="utf-8")
            (question_dir / "answers.jsonl").write_text(json.dumps(answer, ensure_ascii=False) + "\n", encoding="utf-8")
            question_manifest = {
                "item_id": item_id, "state": "complete_visible_qa_exhausted", "total_hint": 1,
                "saved_questions": 1, "saved_answers": 1, "exhausted": True,
                "limit": 0, "limit_reached": False, "finished_at": "2026-08-07T00:00:00+00:00",
            }
            (question_dir / "question_manifest.json").write_text(json.dumps(question_manifest, ensure_ascii=False), encoding="utf-8")
            (root / "data" / "run_manifest.json").write_text(json.dumps({"state": "partial"}), encoding="utf-8")

            summary = build_delivery(root)
            self.assertEqual(summary["products"], 1)
            self.assertEqual(summary["reviews"], 1)
            self.assertEqual(summary["questions"], 1)
            self.assertEqual(summary["answers"], 1)
            self.assertTrue((root / "01_商品资料.xlsx").is_file())
            self.assertTrue((root / "02_评价明细.xlsx").is_file())
            self.assertTrue((root / "03_问大家.xlsx").is_file())
            self.assertTrue((root / "04_采集说明.md").is_file())
            product_book = load_workbook(root / "01_商品资料.xlsx", read_only=True)
            self.assertEqual(product_book.sheetnames, ["商品总览", "规格参数", "SKU快照", "素材索引", "完整性"])
            product_book.close()
            review_book = load_workbook(root / "02_评价明细.xlsx", read_only=True)
            self.assertEqual(review_book.sheetnames, ["评价明细", "采集状态"])
            review_book.close()
            question_book = load_workbook(root / "03_问大家.xlsx", read_only=True)
            self.assertEqual(question_book.sheetnames, ["问题清单", "回答明细", "采集状态"])
            question_book.close()

            package = package_directory(root)
            self.assertTrue(Path(package["zip"]).is_file())
            with zipfile.ZipFile(package["zip"]) as archive:
                names = set(archive.namelist())
            self.assertIn(f"{root.name}/01_商品资料.xlsx", names)
            self.assertFalse(any("profile" in name.lower() or "cookie" in name.lower() for name in names))


if __name__ == "__main__":
    unittest.main()

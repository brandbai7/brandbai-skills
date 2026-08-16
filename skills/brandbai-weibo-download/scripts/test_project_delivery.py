from __future__ import annotations

import json
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from package_delivery import package_directory
from project_delivery import finalize_project_delivery
from project_runner import project_state_dir, run_project_tasks
from test_project_runner import FakeSessionFactory, project_plan


def write_jsonl(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("".join(json.dumps(row, ensure_ascii=False) + "\n" for row in rows), encoding="utf-8")


class ProjectDeliveryTests(unittest.TestCase):
    @staticmethod
    def _executor(*, include_interactions: bool, include_asset: bool):
        def execute(task: dict, **kwargs: object) -> dict:
            task_out = Path(kwargs["task_out"])
            data = task_out / "data"
            data.mkdir(parents=True, exist_ok=True)
            source_type = task["source_type"]
            if source_type == "profile":
                uid = str(task["target_key"])
                write_jsonl(data / "accounts.jsonl", [{
                    "uid": uid, "display_name": f"合成账号{uid[-1]}",
                    "canonical_url": task["canonical_url"],
                    "collected_at": "2026-08-10T00:00:00Z",
                    "completion_state": "complete_visible_account",
                }])
                (data / "profile_selection.json").write_text(json.dumps({
                    "profile_selection_id": f"selection-{uid}", "profile_id": uid,
                    "captured_at": "2026-08-10T00:00:00Z",
                    "state": "complete_visible_pinned_plus_recent_n", "selected": [{
                        "post_id": "AbCdE123", "author_uid": "100001", "author_name": "合成账号1",
                        "rank": 1, "is_pinned": uid == "100001", "selection_reason": "pinned" if uid == "100001" else "recent_non_pinned",
                        "body_preview": "合成官宣微博", "published_at_text": "8月9日 10:00",
                        "canonical_url": "https://weibo.com/100001/AbCdE123",
                    }],
                }, ensure_ascii=False), encoding="utf-8")
                return {"state": "complete_visible_pinned_plus_recent_n", "saved_accounts": 1, "saved_visible_post_refs": 1}
            if source_type == "search":
                write_jsonl(data / "search_snapshots.jsonl", [{
                    "search_snapshot_id": "search-synthetic-1", "query_kind": "keyword",
                    "query": "合成明星 合成品牌", "sort": "综合", "filters": ["全部"],
                    "captured_at": "2026-08-10T00:01:00Z", "state": "complete_first_n_visible_results",
                    "results": [{
                        "rank": 1, "post_id": "AbCdE123", "author_uid": "100001",
                        "author_name": "合成账号1", "body_preview": "合成官宣微博",
                        "canonical_url": "https://weibo.com/100001/AbCdE123",
                    }],
                }])
                return {"state": "complete_first_n_visible_results", "saved_visible_post_refs": 1}
            if source_type == "post":
                write_jsonl(data / "posts.jsonl", [{
                    "post_id": "AbCdE123", "author_uid": "100001", "author_name": "合成账号1",
                    "body": "合成官宣微博正文 #合成话题#", "topics": ["#合成话题#"],
                    "mentions": ["@合成品牌"], "post_type": "image", "published_at_text": "2026-08-09 10:00",
                    "metrics": {"reposts": "3", "comments": "2", "likes": "8"},
                    "canonical_url": "https://weibo.com/100001/AbCdE123",
                    "collected_at": "2026-08-10T00:02:00Z", "completion_state": "complete_visible_post",
                    "completion_note": "合成详情已保存",
                }])
                result = {
                    "state": "complete_visible_test", "post_id": "AbCdE123",
                    "post_state": "complete_visible_post",
                    "comment_state": "not_requested", "repost_state": "not_requested",
                }
                if include_interactions:
                    write_jsonl(data / "comments.jsonl", [{
                        "comment_id": "comment-synthetic-1", "comment_id_type": "platform",
                        "post_id": "AbCdE123", "root_comment_id": "comment-synthetic-1", "level": 1,
                        "author_id": "wb_user_synthetic", "content": "合成评论", "declared_reply_count": 0,
                        "saved_reply_count": 0, "reply_expansion_status": "not_applicable",
                        "collected_at": "2026-08-10T00:03:00Z",
                    }])
                    write_jsonl(data / "reposts.jsonl", [{
                        "repost_id": "repost-synthetic-1", "repost_id_type": "platform",
                        "source_post_id": "AbCdE123", "author_id": "wb_user_synthetic_2",
                        "content": "合成转发", "chain_status": "one_hop_visible",
                        "collected_at": "2026-08-10T00:04:00Z",
                    }])
                    result.update({
                        "comment_state": "complete_visible_comments_exhausted",
                        "repost_state": "complete_visible_reposts_exhausted",
                    })
                if include_asset:
                    local = Path("06_微博素材") / "AbCdE123" / "001_image.jpg"
                    target = task_out / local
                    target.parent.mkdir(parents=True, exist_ok=True)
                    target.write_bytes(b"synthetic-image")
                    write_jsonl(data / "assets.jsonl", [{
                        "asset_id": "weibo:AbCdE123:image:001", "post_id": "AbCdE123",
                        "kind": "image", "order": 1, "requested": True, "status": "downloaded",
                        "local_file": local.as_posix(), "source_url": "https://example.invalid/synthetic.jpg",
                        "bytes": len(b"synthetic-image"), "sha256": "",
                    }])
                return result
            raise AssertionError(source_type)
        return execute

    def test_merges_sources_builds_conditional_files_and_packages_without_state(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path.cwd()) as temp:
            root = Path(temp)
            out = root / "BrandBAI_微博项目普通版"
            plan = project_plan()
            plan["deep_capture"]["download_assets"] = True
            run_project_tasks(
                plan, profile_dir=root / "profile", out=out, mode="all", assets=["images"], resume=False,
                task_executor=self._executor(include_interactions=True, include_asset=True),
                session_factory=FakeSessionFactory(),
            )
            result = finalize_project_delivery(out)
            counts = result["merge"]["counts"]
            self.assertEqual(counts["posts"], 1)
            self.assertEqual(counts["post_sources"], 4)
            self.assertEqual(counts["comments"], 1)
            self.assertEqual(counts["reposts"], 1)
            for name in [
                "00_项目采集总览.xlsx", "01_账号资料.xlsx", "02_微博清单.xlsx",
                "03_评论明细.xlsx", "04_转发扩散.xlsx", "05_搜索与话题快照.xlsx",
                "06_微博素材", "07_采集说明.md", "handoff/analysis_input_manifest.json",
            ]:
                self.assertTrue((out / name).exists(), name)
            workbook = load_workbook(out / "02_微博清单.xlsx", read_only=True)
            try:
                self.assertEqual(workbook["微博详情"]["A2"].value, "AbCdE123")
                self.assertEqual(workbook["可见内容池"].max_row, 5)
                self.assertEqual(workbook["来源关系"].max_row, 5)
            finally:
                workbook.close()
            handoff = json.loads((out / "handoff" / "analysis_input_manifest.json").read_text(encoding="utf-8"))
            self.assertTrue(handoff["structurally_ready_for_analysis"])
            self.assertFalse(handoff["analysis_sufficiency_assessed"])
            package = package_directory(out)
            self.assertTrue(Path(package["sha256_file"]).is_file())
            state_name = project_state_dir(out).name
            with zipfile.ZipFile(package["zip"]) as archive:
                self.assertFalse(any(state_name in name for name in archive.namelist()))

    def test_omits_unrequested_empty_interaction_and_material_outputs(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path.cwd()) as temp:
            root = Path(temp)
            out = root / "delivery"
            plan = project_plan()
            plan["deep_capture"]["download_assets"] = False
            run_project_tasks(
                plan, profile_dir=root / "profile", out=out, mode="posts", assets=[], resume=False,
                task_executor=self._executor(include_interactions=False, include_asset=False),
                session_factory=FakeSessionFactory(),
            )
            finalize_project_delivery(out)
            self.assertFalse((out / "03_评论明细.xlsx").exists())
            self.assertFalse((out / "04_转发扩散.xlsx").exists())
            self.assertFalse((out / "06_微博素材").exists())
            self.assertTrue((out / "05_搜索与话题快照.xlsx").is_file())


if __name__ == "__main__":
    unittest.main()

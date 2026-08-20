"""Build customer-readable workbooks from normalized Weibo JSONL files."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Iterable

from collector_core import atomic_write_json, utc_now


class DeliveryError(RuntimeError):
    pass


def _load_json(path: Path, default: Any) -> Any:
    if not path.is_file():
        return default
    return json.loads(path.read_text(encoding="utf-8-sig"))


def _load_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        return []
    rows: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        if line.strip():
            value = json.loads(line)
            if isinstance(value, dict):
                rows.append(value)
    return rows


def _flatten(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return "；".join(_flatten(item) for item in value if _flatten(item))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return str(value)


def _write_sheet(workbook: Any, title: str, headers: list[str], rows: Iterable[list[Any]]) -> None:
    sheet = workbook.create_sheet(title=title)
    sheet.append(headers)
    for row in rows:
        sheet.append([_flatten(value) for value in row])


def _style(workbook: Any) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor="E6162D")
    font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = font
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in sheet.columns:
            letter = column[0].column_letter
            lengths = [len(str(cell.value or "")) for cell in column[:200]]
            sheet.column_dimensions[letter].width = min(max(max(lengths, default=8) + 2, 10), 48)
        if sheet.max_column == 2:
            sheet.column_dimensions["B"].width = min(max(sheet.column_dimensions["B"].width or 10, 64), 72)


def _save_verified(workbook: Any, path: Path) -> None:
    from openpyxl import load_workbook

    workbook.save(path)
    load_workbook(path, read_only=True).close()


def build_delivery(out_value: str | Path) -> dict[str, Any]:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise DeliveryError("openpyxl is missing") from exc

    out = Path(out_value).expanduser().resolve()
    data = out / "data"
    accounts = _load_jsonl(data / "accounts.jsonl")
    posts = _load_jsonl(data / "posts.jsonl")
    comments = _load_jsonl(data / "comments.jsonl")
    reposts = _load_jsonl(data / "reposts.jsonl")
    assets = _load_jsonl(data / "assets.jsonl")
    searches = _load_jsonl(data / "search_snapshots.jsonl")
    hotlists = _load_jsonl(data / "hotlist_snapshots.jsonl")
    run_manifest = _load_json(data / "run_manifest.json", {})
    profile_selection = _load_json(data / "profile_selection.json", {})
    if not any([accounts, posts, comments, reposts, searches, hotlists]):
        raise DeliveryError("No Weibo collection data was found")
    out.mkdir(parents=True, exist_ok=True)

    account_book = Workbook()
    account_book.remove(account_book.active)
    _write_sheet(account_book, "账号总览", [
        "账号UID", "账号名称", "认证原文", "简介", "关注原文", "粉丝原文", "微博数原文",
        "规范链接", "采集时间", "完成状态",
    ], [[
        row.get("uid"), row.get("display_name"), row.get("verification_text"), row.get("description"),
        row.get("following_text"), row.get("followers_text"), row.get("posts_text"), row.get("canonical_url"),
        row.get("collected_at"), row.get("completion_state"),
    ] for row in accounts])
    _write_sheet(account_book, "主页选择", [
        "主页选择ID", "账号UID", "微博ID", "主页位次", "是否置顶", "选择原因", "发布时间原文", "页面摘要", "规范链接",
    ], [[
        profile_selection.get("profile_selection_id"), profile_selection.get("profile_id"), row.get("post_id"),
        row.get("rank"), row.get("is_pinned"), row.get("selection_reason"), row.get("published_at_text"),
        row.get("body_preview"), row.get("canonical_url"),
    ] for row in profile_selection.get("selected") or []])
    creator_snapshot = posts[0].get("creator_snapshot") if (
        len(posts) == 1 and run_manifest.get("target_kind") == "posts"
        and not profile_selection and not searches and not hotlists
    ) else None
    if isinstance(creator_snapshot, dict) and any(
        creator_snapshot.get(key) not in (None, "")
        for key in ("nickname", "platform_account", "stable_creator_id")
    ):
        _write_sheet(account_book, "达人快照", ["字段", "值"], [
            ["昵称", creator_snapshot.get("nickname")],
            ["微博UID", creator_snapshot.get("platform_account")],
            ["稳定达人ID", creator_snapshot.get("stable_creator_id")],
            ["主页链接", creator_snapshot.get("profile_url")],
            ["简介", creator_snapshot.get("bio")],
            ["粉丝数", creator_snapshot.get("followers")],
            ["累计获赞", creator_snapshot.get("total_likes")],
            ["快照时间", creator_snapshot.get("snapshot_at")],
            ["来源微博ID", posts[0].get("post_id")],
            ["来源微博链接", posts[0].get("canonical_url")],
            ["采集边界", "仅使用当前微博详情页已经展示或加载的公开作者信息；未自动进入账号主页；未下载头像。"],
        ])
    _style(account_book)
    _save_verified(account_book, out / "01_账号资料.xlsx")

    post_book = Workbook()
    post_book.remove(post_book.active)
    _write_sheet(post_book, "微博总览", [
        "微博ID", "账号UID", "作者", "主页UID", "主页位次", "搜索快照ID", "搜索位次", "搜索词", "选择原因",
        "是否置顶", "正文", "类型", "发布时间原文", "地区原文", "来源原文", "可见范围", "浏览", "转发", "评论", "点赞",
        "原微博ID", "原作者UID", "规范链接", "采集时间", "完成状态",
    ], [[
        row.get("post_id"), row.get("author_uid"), row.get("author_name"), row.get("profile_id"), row.get("profile_rank"),
        row.get("search_snapshot_id"), row.get("search_rank"), row.get("search_query"), row.get("selection_reason"),
        row.get("is_pinned"), row.get("body"), row.get("post_type"), row.get("published_at_text"), row.get("region_text"),
        row.get("source_text"), row.get("visibility_text"), (row.get("metrics") or {}).get("views"),
        (row.get("metrics") or {}).get("reposts"), (row.get("metrics") or {}).get("comments"),
        (row.get("metrics") or {}).get("likes"), row.get("original_post_id"), row.get("original_author_uid"),
        row.get("canonical_url"), row.get("collected_at"), row.get("completion_state"),
    ] for row in posts])
    topic_rows = []
    for post in posts:
        for order, topic in enumerate(post.get("topics") or [], start=1):
            topic_rows.append([post.get("post_id"), "topic", order, topic])
        for order, mention in enumerate(post.get("mentions") or [], start=1):
            topic_rows.append([post.get("post_id"), "mention", order, mention])
    _write_sheet(post_book, "话题与提及", ["微博ID", "类型", "顺序", "原文"], topic_rows)
    _write_sheet(post_book, "素材索引", [
        "素材ID", "微博ID", "类型", "顺序", "状态", "本地文件", "来源URL", "宽", "高", "字节数", "SHA256", "失败原因",
    ], [[
        row.get("asset_id"), row.get("post_id"), row.get("kind"), row.get("order"), row.get("status"), row.get("local_file"),
        row.get("source_url"), row.get("width"), row.get("height"), row.get("bytes"), row.get("sha256"), row.get("error_reason"),
    ] for row in assets])
    _write_sheet(post_book, "完整性", ["微博ID", "状态", "说明", "采集时间"], [[
        row.get("post_id"), row.get("completion_state"), row.get("completion_note"), row.get("collected_at")
    ] for row in posts])
    _style(post_book)
    _save_verified(post_book, out / "02_微博清单.xlsx")

    comment_book = Workbook()
    comment_book.remove(comment_book.active)
    _write_sheet(comment_book, "评论明细", [
        "评论ID", "ID类型", "微博ID", "父评论ID", "根评论ID", "层级", "匿名作者ID", "作者显示名", "评论原文",
        "时间原文", "地区原文", "点赞原文", "声明回复数", "已保存回复数", "回复展开状态",
        "观察排序", "各排序首次位次", "采集时间", "最后观察时间",
    ], [[
        row.get("comment_id"), row.get("comment_id_type"), row.get("post_id"), row.get("parent_comment_id"),
        row.get("root_comment_id"), row.get("level"), row.get("author_id"), row.get("author_display"), row.get("content"),
        row.get("time_text"), row.get("region_text"), row.get("like_count_text"), row.get("declared_reply_count"),
        row.get("saved_reply_count"), row.get("reply_expansion_status"),
        "、".join(row.get("observed_sort_modes") or []), json.dumps(row.get("sort_rank_by_mode") or {}, ensure_ascii=False),
        row.get("collected_at"), row.get("last_observed_at"),
    ] for row in comments])
    _write_sheet(comment_book, "采集状态", [
        "微博ID", "保存一级评论", "保存回复", "完成状态", "页面可见排序", "已跑完排序", "终止依据",
    ], _interaction_status_rows(
        comments, "post_id", "level", run_manifest.get("comment_states", {}),
        run_manifest.get("comment_details", {}),
    ))
    _style(comment_book)
    _save_verified(comment_book, out / "03_评论明细.xlsx")

    repost_book = Workbook()
    repost_book.remove(repost_book.active)
    _write_sheet(repost_book, "转发明细", [
        "转发ID", "ID类型", "源微博ID", "上游转发ID", "匿名作者ID", "作者显示名", "转发文案", "时间原文",
        "地区原文", "互动快照", "传播链状态", "采集时间",
    ], [[
        row.get("repost_id"), row.get("repost_id_type"), row.get("source_post_id"), row.get("upstream_repost_id"),
        row.get("author_id"), row.get("author_display"), row.get("content"), row.get("time_text"), row.get("region_text"),
        row.get("metrics"), row.get("chain_status"), row.get("collected_at"),
    ] for row in reposts])
    per_post_reposts: dict[str, int] = {}
    for row in reposts:
        key = str(row.get("source_post_id") or "")
        per_post_reposts[key] = per_post_reposts.get(key, 0) + 1
    repost_states = run_manifest.get("repost_states", {})
    repost_status_ids = sorted(set(per_post_reposts) | set(repost_states))
    _write_sheet(repost_book, "采集状态", ["微博ID", "保存转发记录", "完成状态"], [[
        post_id, per_post_reposts.get(post_id, 0), repost_states.get(post_id, "unknown")
    ] for post_id in repost_status_ids])
    _style(repost_book)
    _save_verified(repost_book, out / "04_转发扩散.xlsx")

    search_book = Workbook()
    search_book.remove(search_book.active)
    result_rows = []
    context_rows = []
    supertopic_rows = []
    hotlist_rows = []
    for snapshot in searches:
        for result in snapshot.get("results") or []:
            result_rows.append([
                snapshot.get("search_snapshot_id"), snapshot.get("query_kind"), snapshot.get("query"), snapshot.get("sort"),
                snapshot.get("filters"), result.get("rank"), result.get("post_id"), result.get("author_uid"),
                result.get("author_name"), result.get("body_preview"), result.get("promoted_state"), result.get("canonical_url"),
                snapshot.get("captured_at"), snapshot.get("state"),
            ])
        context = snapshot.get("topic_context") or {}
        if context or snapshot.get("query_kind") == "topic":
            context_rows.append([
                snapshot.get("search_snapshot_id"), snapshot.get("query"), context.get("read_text"),
                context.get("discuss_text"), context.get("host_text"), snapshot.get("captured_at"),
            ])
        supertopic = snapshot.get("supertopic_context") or {}
        if supertopic or snapshot.get("query_kind") == "supertopic":
            supertopic_rows.append([
                snapshot.get("search_snapshot_id"), supertopic.get("supertopic_id"), supertopic.get("name"),
                supertopic.get("canonical_url"), supertopic.get("category_text"), supertopic.get("post_count_text"),
                supertopic.get("member_count_text"), supertopic.get("member_label_text"), supertopic.get("checkin_text"),
                supertopic.get("rank_text"), supertopic.get("visible_tabs"), supertopic.get("selected_tab") or snapshot.get("sort"),
                snapshot.get("captured_at"), snapshot.get("state"),
            ])
    for snapshot in hotlists:
        for entry in snapshot.get("entries") or []:
            hotlist_rows.append([
                snapshot.get("hotlist_snapshot_id"), snapshot.get("category_code"), snapshot.get("category_name"),
                entry.get("observed_position"), entry.get("rank_text"), entry.get("rank_numeric"),
                entry.get("keyword"), entry.get("heat_text"), entry.get("topic_category_text"), entry.get("label_text"),
                entry.get("is_pinned"), entry.get("is_special"), entry.get("query_url"),
                snapshot.get("captured_at"), snapshot.get("state"),
            ])
    _write_sheet(search_book, "结果位次", [
        "搜索快照ID", "查询类型", "查询词", "排序", "筛选", "位次", "微博ID", "账号UID", "作者", "摘要",
        "推广标记", "规范链接", "采集时间", "范围状态",
    ], result_rows)
    _write_sheet(search_book, "话题上下文", [
        "搜索快照ID", "话题", "阅读原文", "讨论原文", "主持原文", "采集时间",
    ], context_rows)
    _write_sheet(search_book, "超话资料", [
        "搜索快照ID", "超话ID", "超话名称", "规范链接", "分类原文", "帖子数原文", "成员数原文",
        "成员称呼原文", "今日签到原文", "排行原文", "页面可见分区", "采集分区", "采集时间", "范围状态",
    ], supertopic_rows)
    _write_sheet(search_book, "热搜榜单", [
        "榜单快照ID", "榜单代码", "榜单名称", "页面可见顺序", "排名原文", "数字排名", "词条",
        "热度原文", "词条分类原文", "榜单标签", "是否置顶", "是否特殊行", "词条搜索链接", "采集时间", "范围状态",
    ], hotlist_rows)
    _style(search_book)
    _save_verified(search_book, out / "05_搜索与话题快照.xlsx")

    (out / "06_微博素材").mkdir(parents=True, exist_ok=True)
    description = f"""# BrandBAI 微博采集说明

## 本次交付

- 账号记录：{len(accounts)}
- 微博记录：{len(posts)}
- 评论与回复记录：{len(comments)}
- 转发记录：{len(reposts)}
- 搜索与话题快照：{len(searches)}
- 热搜榜单快照：{len(hotlists)}（榜单行 {sum(len(item.get('entries') or []) for item in hotlists)}）
- 运行状态：{run_manifest.get('state', 'unknown')}
- 主页选择状态：{profile_selection.get('state', 'not_applicable')}
- 搜索选择状态：{run_manifest.get('search_selection_state', 'not_applicable')}
- 热搜榜单状态：{run_manifest.get('hotlist_selection_state', 'not_applicable')}
- 构建时间：{utc_now()}

## 重要边界

1. 账号主页范围只对应采集时点页面可见置顶和最近非置顶微博；置顶为额外项，不占最近 N 条名额。
2. 搜索、话题、超话与热搜榜单只对应指定入口、排序、筛选与采集时点的页面可见快照，不代表平台全量或长期排名。
3. 评论和转发完成只表示页面当前可返回列表收到终止信号；回复或传播链未展开时保留部分完成状态。
4. 粉丝、浏览、转发、评论、点赞、话题阅读和讨论均为采集时点快照。
5. 评论、转发和微博正文中的身份、关系、购买、效果和体验主张未经本下载阶段核验。
6. 本包只做下载、结构化和质量说明，不生成明星口碑、粉丝画像、舆情倾向、代言匹配或商业归因。
7. 登录资料、Cookie、请求头、验证码、手机号、登录二维码和带鉴权参数的媒体地址不进入交付包。
"""
    (out / "07_采集说明.md").write_text(description, encoding="utf-8")
    summary = {
        "built_at": utc_now(), "accounts": len(accounts), "posts": len(posts), "comments": len(comments),
        "reposts": len(reposts), "search_snapshots": len(searches), "run_state": run_manifest.get("state", "unknown"),
        "hotlist_snapshots": len(hotlists), "hotlist_entries": sum(len(item.get("entries") or []) for item in hotlists),
        "profile_selection": bool(profile_selection), "profile_selected_posts": len(profile_selection.get("selected") or []),
    }
    atomic_write_json(data / "delivery_manifest.json", summary)
    return summary


def _interaction_status_rows(
    rows: list[dict[str, Any]], object_field: str, level_field: str, states: dict[str, Any],
    details: dict[str, Any] | None = None,
) -> list[list[Any]]:
    per_object: dict[str, dict[str, int]] = {}
    for row in rows:
        key = str(row.get(object_field) or "")
        item = per_object.setdefault(key, {"first": 0, "replies": 0})
        item["first" if int(row.get(level_field) or 1) == 1 else "replies"] += 1
    object_ids = sorted(set(per_object) | set(states))
    output: list[list[Any]] = []
    for key in object_ids:
        detail = dict((details or {}).get(key) or {})
        reasons = [
            str(run.get("termination_reason") or "")
            for run in detail.get("sort_runs") or [] if run.get("termination_reason")
        ]
        output.append([
            key,
            per_object.get(key, {}).get("first", 0),
            per_object.get(key, {}).get("replies", 0),
            states.get(key, "unknown"),
            "、".join(detail.get("sort_modes_available") or []),
            "、".join(detail.get("sort_modes_exhausted") or []),
            "；".join(reasons),
        ])
    return output


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--out", required=True)
    args = parser.parse_args()
    print(json.dumps(build_delivery(args.out), ensure_ascii=False, indent=2))

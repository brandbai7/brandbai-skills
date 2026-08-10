from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from build_delivery import build_delivery
from openpyxl import load_workbook


def write_jsonl(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("".join(json.dumps(row, ensure_ascii=False) + "\n" for row in rows), encoding="utf-8")


class BuildDeliveryTests(unittest.TestCase):
    def test_builds_full_ordinary_delivery_with_zero_row_statuses(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path.cwd()) as temp:
            out = Path(temp) / "delivery"
            data = out / "data"
            data.mkdir(parents=True)
            write_jsonl(data / "accounts.jsonl", [{
                "uid": "10001", "display_name": "合成明星账号", "verification_text": "合成认证",
                "description": "仅用于测试", "canonical_url": "https://weibo.com/u/10001",
                "collected_at": "2026-08-09T00:00:00Z", "completion_state": "complete_visible_profile",
            }])
            write_jsonl(data / "posts.jsonl", [{
                "post_id": "Abcde123", "author_uid": "10001", "author_name": "合成明星账号",
                "body": "合成微博正文 #测试#", "topics": ["#测试#"], "mentions": ["@合成品牌"],
                "post_type": "image", "metrics": {"reposts": "2", "comments": "1", "likes": "3"},
                "canonical_url": "https://weibo.com/10001/Abcde123", "collected_at": "2026-08-09T00:00:00Z",
                "completion_state": "complete_visible_post",
            }])
            write_jsonl(data / "comments.jsonl", [{
                "comment_id": "comment-1", "comment_id_type": "platform", "post_id": "Abcde123",
                "root_comment_id": "comment-1", "level": 1, "author_id": "wb_user_test",
                "content": "合成评论", "declared_reply_count": 0, "saved_reply_count": 0,
                "reply_expansion_status": "not_applicable",
            }])
            write_jsonl(data / "search_snapshots.jsonl", [{
                "search_snapshot_id": "search-1", "query_kind": "keyword", "query": "合成品牌",
                "sort": "综合", "filters": ["全部"], "captured_at": "2026-08-09T00:00:00Z",
                "state": "complete_first_n_visible_results", "results": [{
                    "rank": 1, "post_id": "Abcde123", "author_uid": "10001", "author_name": "合成明星账号",
                    "body_preview": "合成微博正文", "canonical_url": "https://weibo.com/10001/Abcde123",
                }],
            }, {
                "search_snapshot_id": "search-2", "query_kind": "supertopic", "query": "合成明星",
                "sort": "热门", "filters": ["supertopic_id:100808synthetic123"],
                "captured_at": "2026-08-09T00:00:00Z", "state": "complete_first_n_visible_supertopic_posts",
                "supertopic_context": {
                    "supertopic_id": "100808synthetic123", "name": "合成明星",
                    "canonical_url": "https://weibo.com/p/100808synthetic123/super_index?mod=TAB",
                    "category_text": "明星超话", "post_count_text": "100万", "member_count_text": "10万",
                    "member_label_text": "合成人", "checkin_text": "1万", "rank_text": "No.1",
                    "visible_tabs": ["热门", "最新", "精华"], "selected_tab": "热门",
                },
                "results": [{
                    "rank": 1, "post_id": "Abcde123", "author_uid": "10001", "author_name": "合成明星账号",
                    "body_preview": "合成超话微博", "canonical_url": "https://weibo.com/10001/Abcde123",
                }],
            }])
            write_jsonl(data / "hotlist_snapshots.jsonl", [{
                "hotlist_snapshot_id": "hotlist-1", "category_code": "entrank", "category_name": "文娱",
                "captured_at": "2026-08-09T00:00:00Z", "state": "complete_ranked_hotlist_plus_visible_extras",
                "entries": [{
                    "observed_position": 1, "rank_text": "1", "rank_numeric": 1,
                    "keyword": "合成明星新剧", "heat_text": "100万", "topic_category_text": "电视剧",
                    "label_text": "沸", "is_pinned": False, "is_special": False,
                    "query_url": "https://s.weibo.com/weibo?q=%23%E5%90%88%E6%88%90%23",
                }],
            }])
            (data / "run_manifest.json").write_text(json.dumps({
                "state": "partial", "comment_states": {"Abcde123": "complete_visible_comments_exhausted"},
                "repost_states": {"Abcde123": "partial_reposts_not_available"},
            }), encoding="utf-8")
            (data / "profile_selection.json").write_text(json.dumps({
                "profile_selection_id": "selection-1", "profile_id": "10001",
                "state": "complete_visible_pinned_plus_recent_n", "selected": [{
                    "post_id": "Abcde123", "rank": 1, "is_pinned": True,
                    "selection_reason": "pinned", "canonical_url": "https://weibo.com/10001/Abcde123",
                }],
            }), encoding="utf-8")

            summary = build_delivery(out)
            self.assertEqual(summary["posts"], 1)
            self.assertEqual(summary["comments"], 1)
            self.assertEqual(summary["hotlist_entries"], 1)
            for name in [
                "01_账号资料.xlsx", "02_微博清单.xlsx", "03_评论明细.xlsx",
                "04_转发扩散.xlsx", "05_搜索与话题快照.xlsx", "07_采集说明.md",
            ]:
                self.assertTrue((out / name).exists(), name)
            workbook = load_workbook(out / "04_转发扩散.xlsx", read_only=True)
            try:
                self.assertEqual(workbook["采集状态"]["A2"].value, "Abcde123")
                self.assertEqual(workbook["采集状态"]["B2"].value, "0")
                self.assertEqual(workbook["采集状态"]["C2"].value, "partial_reposts_not_available")
            finally:
                workbook.close()
            workbook = load_workbook(out / "05_搜索与话题快照.xlsx", read_only=True)
            try:
                self.assertEqual(workbook["超话资料"]["B2"].value, "100808synthetic123")
                self.assertEqual(workbook["超话资料"]["L2"].value, "热门")
                self.assertEqual(workbook["热搜榜单"]["G2"].value, "合成明星新剧")
                self.assertEqual(workbook["热搜榜单"]["J2"].value, "沸")
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()

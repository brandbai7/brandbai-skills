import csv
import json
import shutil
import unittest
import uuid
from contextlib import contextmanager
from pathlib import Path

from openpyxl import load_workbook

from build_foundation_workbooks import main


@contextmanager
def workspace_temp():
    root = Path.cwd() / "_foundation_test_artifacts"
    root.mkdir(parents=True, exist_ok=True)
    path = root / f"workbook_{uuid.uuid4().hex}"
    path.mkdir()
    try:
        yield path
    finally:
        shutil.rmtree(path, ignore_errors=True)
        try:
            root.rmdir()
        except OSError:
            pass


class BuildFoundationWorkbooksTests(unittest.TestCase):
    def test_builds_portable_ordinary_delivery(self):
        with workspace_temp() as temp:
            works_path = temp / "works.json"
            works_manifest_path = temp / "download_manifest.json"
            comments_path = temp / "comments.csv"
            comments_manifest_path = temp / "run_manifest.json"
            output_dir = temp / "delivery"
            qa_dir = temp / "qa"

            works_path.write_text(json.dumps({"works": [{
                "aweme_id": "7551794579813502262",
                "type": "视频",
                "author": "测试达人",
                "title": "这是一条带标题的测试作品",
                "publish_time": "2026-08-01T10:30:00+08:00",
                "digg_count": 123,
                "comment_count": 2,
                "collect_count": 9,
                "share_count": 4,
                "is_pinned": "0",
                "selection_reason": "最近",
                "source_url": "https://www.douyin.com/video/7551794579813502262",
                "local_folder": "media/作品_7551794579813502262",
                "download_status": "完成",
                "downloads": {
                    "video": {"file": "video.mp4", "status": "downloaded", "bytes": 1048576},
                    "cover": {"file": "cover.jpg", "status": "downloaded", "bytes": 2048},
                    "music": {"status": "not_available"},
                },
            }]} , ensure_ascii=False), encoding="utf-8")
            works_manifest_path.write_text(json.dumps({
                "visible_works_observed": 12,
                "requested_recent_non_pinned": 5,
                "status": "complete",
                "finished_at": "2026-08-01T12:00:00+08:00",
            }), encoding="utf-8")

            headers = [
                "aweme_id", "comment_id", "root_comment_id", "parent_comment_id", "reply_level",
                "text", "author_pseudonym", "create_time", "digg_count", "reply_count",
                "source_role", "source_url", "ip_label", "is_pinned", "is_creator_reply",
                "evidence_state", "evidence_id", "collected_at",
            ]
            with comments_path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=headers)
                writer.writeheader()
                writer.writerow({
                    "aweme_id": "7551794579813502262",
                    "comment_id": "7551794579813502262001",
                    "root_comment_id": "7551794579813502262001",
                    "reply_level": "0",
                    "text": "这条评论用于验证普通版导出。",
                    "author_pseudonym": "用户_001",
                    "create_time": "2026-08-01T11:00:00+08:00",
                    "digg_count": "8",
                    "reply_count": "0",
                    "source_role": "viewer_comment",
                    "source_url": "https://www.douyin.com/video/7551794579813502262",
                    "ip_label": "上海",
                    "is_pinned": "0",
                    "is_creator_reply": "0",
                    "evidence_state": "F",
                    "evidence_id": "DY-C-7551794579813502262001",
                    "collected_at": "2026-08-01T12:00:00+08:00",
                })
            comments_manifest_path.write_text(json.dumps({
                "include_replies": False,
                "status": "complete_source_visible",
                "privacy_mode": "hash",
                "worker_page_retries": 0,
                "worker_page_crashes": 0,
                "started_at": "2026-08-01T11:55:00+08:00",
                "finished_at": "2026-08-01T12:00:00+08:00",
            }), encoding="utf-8")

            result = main([
                str(works_path), str(works_manifest_path), str(comments_path),
                str(comments_manifest_path), str(output_dir), "--qa-dir", str(qa_dir),
            ])
            self.assertEqual(result, 0)
            self.assertTrue((output_dir / "04_采集说明.md").is_file())
            self.assertTrue((qa_dir / "workbook_qa.json").is_file())

            works_book = load_workbook(output_dir / "01_作品清单.xlsx")
            self.assertEqual(works_book.sheetnames, ["使用说明", "作品清单", "素材明细"])
            self.assertEqual(works_book["作品清单"]["A2"].value, "7551794579813502262")
            self.assertEqual(works_book["作品清单"]["A2"].number_format, "@")
            self.assertTrue(works_book["作品清单"]["A2"].quotePrefix)
            self.assertEqual(works_book["作品清单"]["E2"].value.hour, 10)
            self.assertEqual(works_book["使用说明"]["B6"].value, 0)
            self.assertEqual(works_book["使用说明"]["B21"].value, "完成")
            self.assertIn("WorksTable", works_book["作品清单"].tables)
            self.assertEqual(
                works_book["作品清单"]["M2"].hyperlink.target,
                "https://www.douyin.com/video/7551794579813502262",
            )
            self.assertIn("03_%E4%BD%9C%E5%93%81%E7%B4%A0%E6%9D%90", works_book["作品清单"]["N2"].hyperlink.target)
            self.assertIsNotNone(works_book["素材明细"]["I2"].hyperlink)
            works_book.close()

            comments_book = load_workbook(output_dir / "02_评论明细.xlsx")
            self.assertEqual(comments_book.sheetnames, [
                "导出说明", "视频清单", "DataTool兼容", "评论明细", "采集质量", "字段字典",
            ])
            self.assertEqual(comments_book["评论明细"]["J2"].value, "7551794579813502262001")
            self.assertEqual(comments_book["评论明细"]["J2"].number_format, "@")
            self.assertTrue(comments_book["评论明细"]["J2"].quotePrefix)
            self.assertEqual(comments_book["评论明细"]["D2"].value.hour, 11)
            self.assertEqual(comments_book["DataTool兼容"]["A2"].value, "这条评论用于验证普通版导出。")
            self.assertIn("DataToolView", comments_book["DataTool兼容"].tables)
            self.assertEqual(comments_book["评论明细"].freeze_panes, "C2")
            self.assertIsNotNone(comments_book["评论明细"]["H2"].hyperlink)
            self.assertIsNotNone(comments_book["视频清单"]["E2"].hyperlink)
            self.assertIn("静态查看快照", comments_book["导出说明"]["A17"].value)
            self.assertEqual(comments_book["导出说明"]["B8"].value, "一级评论完整")
            self.assertEqual(comments_book["采集质量"]["G2"].value, "一级评论完整")
            self.assertIn("CommentsTable", comments_book["评论明细"].tables)
            comments_book.close()

            qa = json.loads((qa_dir / "workbook_qa.json").read_text(encoding="utf-8"))
            self.assertEqual(qa["01_作品清单.xlsx"]["formula_errors"], [])
            self.assertEqual(qa["02_评论明细.xlsx"]["formula_errors"], [])
            self.assertEqual(qa["01_作品清单.xlsx"]["id_text_errors"], [])
            self.assertEqual(qa["02_评论明细.xlsx"]["id_text_errors"], [])
            self.assertGreater(qa["01_作品清单.xlsx"]["sheets"]["作品清单"]["hyperlinks"], 0)
            self.assertGreater(qa["02_评论明细.xlsx"]["sheets"]["评论明细"]["hyperlinks"], 0)


if __name__ == "__main__":
    unittest.main()

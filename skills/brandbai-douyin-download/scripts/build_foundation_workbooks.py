#!/usr/bin/env python3
"""Build portable BrandBAI ordinary-delivery workbooks with openpyxl."""

from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path, PureWindowsPath
from typing import Any, Iterable
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


NAVY = "17324D"
TEAL = "159A9C"
PALE_TEAL = "EAF6F5"
PALE_BLUE = "EDF3F8"
PALE_AMBER = "FFF5D6"
GREEN = "E7F6ED"
GREEN_TEXT = "196B3A"
GRAY = "5B6573"
DARK = "24313D"
BORDER = "D7E1E8"
WHITE = "FFFFFF"
LINK_BLUE = "0563C1"
FONT_NAME = "Microsoft YaHei"
THIN = Side(style="thin", color=BORDER)
CHINA_TZ = timezone(timedelta(hours=8))


def configure_output() -> None:
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            try:
                stream.reconfigure(encoding="utf-8", errors="replace")
            except (AttributeError, ValueError):
                pass


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def load_works(path: Path) -> list[dict[str, Any]]:
    payload = load_json(path)
    rows = payload.get("works") if isinstance(payload, dict) else payload
    if not isinstance(rows, list):
        raise ValueError("works.json must be a list or contain a works list")
    return [row for row in rows if isinstance(row, dict)]


def load_comments(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def as_number(value: Any, default: int | float | None = None) -> int | float | None:
    if value is None or value == "":
        return default
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    return int(number) if number.is_integer() else number


def as_datetime(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        parsed = value
    else:
        text = str(value).strip().replace("Z", "+00:00")
        try:
            parsed = datetime.fromisoformat(text)
        except ValueError:
            return None
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(CHINA_TZ).replace(tzinfo=None)
    return parsed


def as_datetime_or_visible_text(value: Any) -> datetime | str | None:
    """Keep page-visible relative times when they cannot be parsed as absolute dates."""
    parsed = as_datetime(value)
    if parsed is not None:
        return parsed
    text = str(value or "").strip()
    return text or None


def id_source(row: dict[str, Any]) -> str:
    explicit = str(row.get("id_source") or "").strip()
    if explicit in {"platform", "dom_fallback"}:
        return explicit
    comment_id = str(row.get("comment_id") or "")
    return "dom_fallback" if comment_id.startswith("generated_") else "platform"


def yes_no(value: Any) -> str:
    return "是" if value is True or str(value).strip().lower() in {"1", "true", "yes"} else "否"


def is_true(value: Any) -> bool:
    return value is True or str(value).strip().lower() in {"1", "true", "yes"}


def display_status(value: Any) -> str:
    raw = str(value or "").strip()
    return {
        "complete": "完成",
        "complete_source_visible": "一级评论完整",
        "not_requested": "未采集",
        "partial_limit_sample": "限额样本",
        "partial": "部分完成",
        "partial_source_visible": "部分完成",
        "failed": "失败",
        "running": "采集中",
    }.get(raw, raw)


def source_role(value: Any) -> str:
    return {
        "viewer_comment": "用户评论",
        "viewer_reply": "用户回复",
        "creator_reply": "作者回复",
    }.get(str(value or ""), str(value or ""))


def asset_status(item: Any) -> str:
    status = str(item.get("status") or "") if isinstance(item, dict) else ""
    return {
        "downloaded": "已下载",
        "skipped_existing": "已存在",
        "created": "已生成",
        "not_requested": "未选择",
        "not_available": "源未提供",
        "failed": "下载失败",
    }.get(status, status)


def set_dimensions(ws: Any, widths: Iterable[float]) -> None:
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def style_header(ws: Any, columns: int) -> None:
    for cell in ws[1][:columns]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30


def style_data_sheet(
    ws: Any,
    columns: int,
    rows: int,
    *,
    widths: Iterable[float],
    wrap_columns: Iterable[int] = (),
    date_columns: Iterable[int] = (),
    integer_columns: Iterable[int] = (),
    decimal_columns: Iterable[int] = (),
    text_columns: Iterable[int] = (),
    row_height: float | None = None,
    table_name: str | None = None,
    table_style: str = "TableStyleMedium2",
    freeze_panes: str = "A2",
) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze_panes
    style_header(ws, columns)
    set_dimensions(ws, widths)
    wrap_set = set(wrap_columns)
    date_set = set(date_columns)
    integer_set = set(integer_columns)
    decimal_set = set(decimal_columns)
    text_set = set(text_columns)
    for row_index in range(2, rows + 2):
        if row_height:
            ws.row_dimensions[row_index].height = row_height
        for column_index in range(1, columns + 1):
            cell = ws.cell(row_index, column_index)
            cell.font = Font(name=FONT_NAME, size=10, color=DARK)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=column_index in wrap_set,
            )
            cell.border = Border(bottom=THIN)
            if column_index in date_set:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            elif column_index in integer_set:
                cell.number_format = "#,##0"
            elif column_index in decimal_set:
                cell.number_format = "0.00"
            elif column_index in text_set:
                cell.number_format = "@"
                cell.quotePrefix = True
    if rows and table_name:
        table = Table(displayName=table_name, ref=f"A1:{get_column_letter(columns)}{rows + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name=table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)


def style_hyperlink_column(
    ws: Any,
    rows: int,
    column: int,
    *,
    relative_local_path: bool = False,
) -> None:
    """Make URL or relative-path cells clickable without changing their display values."""
    for row_index in range(2, rows + 2):
        cell = ws.cell(row_index, column)
        value = str(cell.value or "").strip()
        if not value:
            continue
        target = value
        if relative_local_path:
            target = quote(value.replace("\\", "/"), safe="/:")
        cell.hyperlink = target
        cell.font = Font(name=FONT_NAME, size=10, color=LINK_BLUE, underline="single")


def style_title(ws: Any, title: str) -> None:
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = title
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name=FONT_NAME, size=18, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 44


def style_summary(ws: Any, start_row: int, labels: list[str], values: list[Any]) -> None:
    ws.cell(start_row, 1, "任务摘要")
    ws.cell(start_row, 2, "结果")
    for cell in ws[start_row][0:2]:
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(name=FONT_NAME, bold=True, color=WHITE)
    for offset, (label, value) in enumerate(zip(labels, values), start=1):
        row = start_row + offset
        left = ws.cell(row, 1, label)
        right = ws.cell(row, 2, value)
        left.fill = PatternFill("solid", fgColor=PALE_BLUE)
        left.font = Font(name=FONT_NAME, bold=True, color=NAVY)
        right.fill = PatternFill("solid", fgColor=PALE_TEAL)
        right.font = Font(name=FONT_NAME, size=12, bold=True, color=TEAL)
        right.number_format = "#,##0" if isinstance(value, (int, float)) else "General"
    for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(labels), min_col=1, max_col=2):
        for cell in row:
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_notes(ws: Any, heading_row: int, notes: list[str]) -> None:
    ws.merge_cells(start_row=heading_row, start_column=1, end_row=heading_row, end_column=6)
    heading = ws.cell(heading_row, 1, "口径与限制")
    heading.fill = PatternFill("solid", fgColor=PALE_AMBER)
    heading.font = Font(name=FONT_NAME, bold=True, color=NAVY)
    for offset, note in enumerate(notes, start=1):
        row = heading_row + offset
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row, 1, note)
        cell.font = Font(name=FONT_NAME, size=10, color=GRAY)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)
        ws.row_dimensions[row].height = 42


def style_metadata(ws: Any, start_row: int, rows: list[tuple[str, Any]]) -> None:
    for offset, (label, value) in enumerate(rows):
        row = start_row + offset
        left = ws.cell(row, 1, label)
        right = ws.cell(row, 2, value)
        left.fill = PatternFill("solid", fgColor=PALE_BLUE)
        left.font = Font(name=FONT_NAME, bold=True, color=NAVY)
        if isinstance(value, datetime):
            right.number_format = "yyyy-mm-dd hh:mm:ss"


def workbook_properties(book: Workbook, title: str) -> None:
    book.properties.creator = "布兰德老白 BrandBAI"
    book.properties.lastModifiedBy = "布兰德老白 BrandBAI"
    book.properties.title = title
    book.properties.subject = "Douyin public-page collection delivery"
    book.calculation.calcMode = "auto"
    book.calculation.fullCalcOnLoad = True
    book.calculation.forceFullCalc = True


def build_works_book(
    works: list[dict[str, Any]], works_manifest: dict[str, Any], output_path: Path
) -> tuple[int, int, int]:
    creator = str(works[0].get("author") or "未知达人") if works else "未知达人"
    book = Workbook()
    intro = book.active
    intro.title = "使用说明"
    work_list = book.create_sheet("作品清单")
    asset_list = book.create_sheet("素材明细")
    workbook_properties(book, f"BrandBAI 抖音作品采集｜{creator}")

    headers = [
        "作品ID", "类型", "达人", "标题", "发布时间", "点赞数", "评论数", "收藏数", "分享数", "推荐数",
        "是否置顶", "入选口径", "作品链接", "素材文件夹", "下载状态", "原声状态", "发布文案状态",
    ]
    work_list.append(headers)
    for work in works:
        work_list.append([
            str(work.get("aweme_id") or ""), work.get("type") or "", work.get("author") or "",
            work.get("title") or "", as_datetime(work.get("publish_time")),
            as_number(work.get("digg_count"), 0), as_number(work.get("comment_count"), 0),
            as_number(work.get("collect_count"), 0), as_number(work.get("share_count"), 0),
            as_number(work.get("recommend_count")), yes_no(work.get("is_pinned")),
            work.get("selection_reason") or "", work.get("source_url") or "",
            str(work.get("local_folder") or "").replace("media/", "03_作品素材\\").replace("media\\", "03_作品素材\\"),
            work.get("download_status") or "", asset_status((work.get("downloads") or {}).get("music")),
            asset_status((work.get("downloads") or {}).get("caption")),
        ])
    style_data_sheet(
        work_list, 17, len(works), widths=[22, 9, 14, 54, 20, 11, 11, 11, 11, 11, 10, 12, 42, 38, 12, 16, 16],
        wrap_columns=[4, 13, 14, 15, 16, 17], date_columns=[5], integer_columns=[6, 7, 8, 9, 10],
        text_columns=[1], row_height=48, table_name="WorksTable", table_style="TableStyleMedium2",
    )
    style_hyperlink_column(work_list, len(works), 13)
    style_hyperlink_column(work_list, len(works), 14, relative_local_path=True)

    asset_headers = ["作品ID", "标题", "资产类型", "序号", "文件名", "状态", "字节数", "大小(MB)", "本地相对路径"]
    asset_list.append(asset_headers)
    asset_record_count = 0
    media_file_count = 0
    for work in works:
        downloads = work.get("downloads") if isinstance(work.get("downloads"), dict) else {}
        entries: list[tuple[str, int, dict[str, Any]]] = []
        images = downloads.get("images") if isinstance(downloads, dict) else None
        if isinstance(images, list):
            entries.extend(("图文图片", index, item) for index, item in enumerate(images, start=1) if isinstance(item, dict))
        for label, key in (("视频", "video"), ("封面", "cover"), ("原声", "music"), ("发布文案", "caption")):
            item = downloads.get(key) if isinstance(downloads, dict) else None
            if isinstance(item, dict):
                entries.append((label, 1, item))
        folder = str(work.get("local_folder") or "").replace("media/", "03_作品素材\\").replace("media\\", "03_作品素材\\")
        for kind, index, item in entries:
            file_name = str(item.get("file") or "")
            relative = str(PureWindowsPath(folder, file_name)) if file_name else ""
            byte_count = int(as_number(item.get("bytes"), 0) or 0)
            asset_list.append([
                str(work.get("aweme_id") or ""), work.get("title") or "", kind, index, file_name,
                asset_status(item), byte_count, round(byte_count / 1024 / 1024, 2), relative,
            ])
            asset_record_count += 1
            if file_name and str(item.get("status") or "") in {"downloaded", "skipped_existing", "created"}:
                media_file_count += 1
    style_data_sheet(
        asset_list, 9, asset_record_count, widths=[22, 52, 12, 8, 18, 12, 14, 12, 54],
        wrap_columns=[2, 9], integer_columns=[7], decimal_columns=[8], text_columns=[1],
        table_name="AssetsTable", table_style="TableStyleMedium4",
    )
    style_hyperlink_column(asset_list, asset_record_count, 9, relative_local_path=True)

    intro.sheet_view.showGridLines = False
    style_title(intro, f"BrandBAI 抖音作品采集｜{creator}")
    style_summary(
        intro, 3,
        ["主页可见作品", "入选作品", "置顶作品", "最近作品", "完成作品", "视频作品", "图文作品", "素材文件"],
        [
            int(as_number(works_manifest.get("visible_works_observed"), 0) or 0), len(works),
            sum(1 for row in works if is_true(row.get("is_pinned"))),
            sum(1 for row in works if row.get("selection_reason") == "最近"),
            sum(1 for row in works if row.get("download_status") == "完成"),
            sum(1 for row in works if row.get("type") == "视频"),
            sum(1 for row in works if row.get("type") == "图文"), media_file_count,
        ],
    )
    style_notes(intro, 13, [
        f"选样规则：主页全部可见置顶作品，另加最近 {int(as_number(works_manifest.get('requested_recent_non_pinned'), 0) or 0)} 条非置顶作品；置顶不占最近 N 条名额。",
        "素材规则：视频保存视频、封面与公开原声；图文保存全部可见图片、封面与公开原声。",
        "公开页面未提供或明确不可用的附件记为“源未提供”，不绕过平台限制。",
        "互动数据是采集时点快照，之后可能继续变化。",
        "本文件只呈现采集结果，不包含达人分析、语义标签或商业结论。",
    ])
    intro["A13"] = "本次口径与边界"
    style_metadata(intro, 20, [
        ("采集来源", "抖音普通登录页面可见返回"),
        ("任务状态", display_status(works_manifest.get("status"))),
        ("完成时间", as_datetime(works_manifest.get("finished_at"))),
    ])
    set_dimensions(intro, [22, 28, 18, 18, 18, 18])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    book.save(output_path)
    return len(works), asset_record_count, media_file_count


def build_comments_book(
    works: list[dict[str, Any]],
    comments: list[dict[str, str]],
    comments_manifest: dict[str, Any],
    output_path: Path,
) -> tuple[int, int]:
    creator = str(works[0].get("author") or "未知达人") if works else "未知达人"
    work_by_id = {str(work.get("aweme_id") or ""): work for work in works}
    comments_by_work: dict[str, list[dict[str, str]]] = {}
    for row in comments:
        comments_by_work.setdefault(str(row.get("aweme_id") or ""), []).append(row)
    include_replies = bool(comments_manifest.get("include_replies"))
    completion = "完整" if comments_manifest.get("status") == "complete_source_visible" else "部分完成"

    book = Workbook()
    intro = book.active
    intro.title = "导出说明"
    video_list = book.create_sheet("视频清单")
    datatool = book.create_sheet("DataTool兼容")
    comment_list = book.create_sheet("评论明细")
    quality = book.create_sheet("采集质量")
    dictionary = book.create_sheet("字段字典")
    workbook_properties(book, f"BrandBAI 抖音评论采集｜{creator}")

    detail_headers = [
        "作品标题", "评论内容", "评论人", "评论时间", "点赞数", "回复数", "评论角色", "作品链接", "作品ID",
        "评论ID", "ID来源", "根评论ID", "父评论ID", "层级", "IP属地", "是否置顶", "是否作者回复", "证据状态", "证据ID", "采集时间",
    ]
    comment_list.append(detail_headers)
    detail_rows: list[list[Any]] = []
    for comment in comments:
        work = work_by_id.get(str(comment.get("aweme_id") or ""), {})
        row = [
            work.get("title") or "", comment.get("text") or "", comment.get("author_pseudonym") or "",
            as_datetime_or_visible_text(comment.get("create_time")), as_number(comment.get("digg_count"), 0),
            as_number(comment.get("reply_count"), 0), source_role(comment.get("source_role")),
            comment.get("source_url") or work.get("source_url") or "", str(comment.get("aweme_id") or ""),
            str(comment.get("comment_id") or ""), id_source(comment), str(comment.get("root_comment_id") or ""),
            str(comment.get("parent_comment_id") or ""), as_number(comment.get("reply_level"), 0),
            comment.get("ip_label") or "", yes_no(comment.get("is_pinned")),
            yes_no(comment.get("is_creator_reply")), comment.get("evidence_state") or "",
            str(comment.get("evidence_id") or ""), as_datetime(comment.get("collected_at")),
        ]
        comment_list.append(row)
        detail_rows.append(row)
    style_data_sheet(
        comment_list, 20, len(detail_rows), widths=[48, 58, 18, 20, 11, 11, 12, 42, 22, 22, 14, 22, 22, 8, 12, 10, 12, 10, 30, 20],
        wrap_columns=[1, 2, 8], date_columns=[4, 20], integer_columns=[5, 6, 14],
        text_columns=[9, 10, 11, 12, 13, 19], row_height=34, table_name="CommentsTable",
        freeze_panes="C2",
    )
    style_hyperlink_column(comment_list, len(detail_rows), 8)

    video_headers = ["作品ID", "类型", "标题", "发布时间", "作品链接", "页面评论数", "一级已采集", "回复已采集", "完整性", "备注"]
    video_list.append(video_headers)
    video_rows: list[list[Any]] = []
    for work in works:
        rows = comments_by_work.get(str(work.get("aweme_id") or ""), [])
        top = sum(1 for row in rows if int(as_number(row.get("reply_level"), 0) or 0) == 0)
        replies = len(rows) - top
        row = [
            str(work.get("aweme_id") or ""), work.get("type") or "", work.get("title") or "",
            as_datetime(work.get("publish_time")), work.get("source_url") or "",
            as_number(work.get("comment_count")), top, replies, completion,
            "包含可检索回复" if include_replies else "本轮未采二级回复",
        ]
        video_list.append(row)
        video_rows.append(row)
    style_data_sheet(
        video_list, 10, len(video_rows), widths=[22, 9, 54, 20, 42, 13, 13, 13, 12, 22],
        wrap_columns=[3, 5, 10], date_columns=[4], integer_columns=[6, 7, 8], text_columns=[1],
        row_height=42, table_name="VideosTable", table_style="TableStyleMedium4",
    )
    style_hyperlink_column(video_list, len(video_rows), 5)

    datatool.append(["评论内容", "评论人", "评论时间", "点赞数", "回复数"])
    for row in detail_rows:
        datatool.append([row[1], row[2], row[3], row[4], row[5]])
    style_data_sheet(
        datatool, 5, len(detail_rows), widths=[64, 20, 20, 12, 12], wrap_columns=[1],
        date_columns=[3], integer_columns=[4, 5], row_height=34,
        table_name="DataToolView", table_style="TableStyleMedium9",
    )
    datatool.sheet_properties.tabColor = TEAL

    quality.append([
        "作品ID", "标题", "一级分页", "一级评论", "回复", "平台ID评论", "页面兜底评论",
        "是否请求回复", "整体状态", "重试", "崩溃", "备注",
    ])
    quality_rows: list[list[Any]] = []
    for work in works:
        rows = comments_by_work.get(str(work.get("aweme_id") or ""), [])
        top = sum(1 for row in rows if int(as_number(row.get("reply_level"), 0) or 0) == 0)
        replies = len(rows) - top
        platform_ids = sum(1 for item in rows if id_source(item) == "platform")
        dom_fallback_ids = len(rows) - platform_ids
        row = [
            str(work.get("aweme_id") or ""), work.get("title") or "",
            "已终止" if completion == "完整" else "未完全验证", top, replies,
            platform_ids, dom_fallback_ids, "是" if include_replies else "否",
            display_status(comments_manifest.get("status")),
            int(as_number(comments_manifest.get("worker_page_retries"), 0) or 0),
            int(as_number(comments_manifest.get("worker_page_crashes"), 0) or 0),
            "按回复完成状态判断" if include_replies else "只验收一级评论",
        ]
        quality.append(row)
        quality_rows.append(row)
    style_data_sheet(
        quality, 12, len(quality_rows), widths=[22, 52, 14, 12, 12, 13, 14, 14, 28, 10, 10, 24],
        wrap_columns=[2, 9, 12], integer_columns=[4, 5, 6, 7, 10, 11], text_columns=[1],
    )
    if quality_rows:
        quality.conditional_formatting.add(
            f"C2:C{len(quality_rows) + 1}",
            FormulaRule(
                formula=['ISNUMBER(SEARCH("已终止",C2))'],
                fill=PatternFill("solid", fgColor=GREEN),
                font=Font(name=FONT_NAME, bold=True, color=GREEN_TEXT),
            ),
        )

    dictionary_rows = [
        ["字段", "类型", "含义", "空值口径"],
        ["作品ID/评论ID/证据ID", "文本", "平台或证据追踪标识", "未知时留空，不转科学计数法"],
        ["ID来源", "分类", "platform 为平台 ID，dom_fallback 为页面可见卡片生成的兜底 ID", "未知时按评论ID规则推断"],
        ["评论内容", "文本", "本次页面可见的原始评论文字", "空文本保留"],
        ["评论时间/采集时间", "日期时间或可见文本", "平台时间与本次采集时间", "相对时间保留页面原文；未知时留空"],
        ["点赞数/回复数", "整数", "页面返回的互动数量；回复数是该评论声明的子回复数量，不代表已采集回复", "确认无互动为 0，未知时留空"],
        ["层级", "整数", "0 为一级评论，1 为回复", "未知时留空"],
        ["评论角色", "分类", "用户评论、用户回复或作者回复", "未知时留空"],
        ["证据状态", "分类", "评论文字存在属于可观察事实 F", "不代表评论内主张已核验"],
    ]
    for row in dictionary_rows:
        dictionary.append(row)
    style_data_sheet(
        dictionary, 4, len(dictionary_rows) - 1, widths=[30, 16, 58, 42], wrap_columns=[1, 2, 3, 4],
        row_height=42,
    )

    top_level = sum(1 for row in comments if int(as_number(row.get("reply_level"), 0) or 0) == 0)
    replies = len(comments) - top_level
    dom_fallback_total = sum(1 for row in comments if id_source(row) == "dom_fallback")
    intro.sheet_view.showGridLines = False
    style_title(intro, f"BrandBAI 抖音评论采集｜{creator}")
    style_summary(
        intro, 3,
        ["入选作品", "一级评论", "二级回复", "页面兜底评论", "完成作品", "任务状态", "页面重试", "页面崩溃"],
        [
            len(video_rows), top_level, replies, dom_fallback_total,
            sum(1 for row in video_rows if row[8] == "完整"), display_status(comments_manifest.get("status")),
            int(as_number(comments_manifest.get("worker_page_retries"), 0) or 0),
            int(as_number(comments_manifest.get("worker_page_crashes"), 0) or 0),
        ],
    )
    style_notes(intro, 12, [
        "“全部评论”指采集时点普通登录页面能够分页返回、且本次收到终止信号的全部可检索评论，不代表平台内部绝对全量。",
        "本次请求二级回复；整体完整性同时受回复楼层完成状态约束；页面评论数可能同时包含一级评论和回复。" if include_replies else "本次只采一级评论；页面评论数可能同时包含一级评论和其下回复，评论明细中的回复数不代表已采二级回复。",
        "普通评论者默认使用稳定化名；作品、评论和证据 ID 保留用于去重与回溯。",
        "ID来源=platform 表示平台评论 ID；ID来源=dom_fallback 表示页面可见卡片生成的稳定兜底 ID，回溯强度较低。",
        "评论文字的存在可作为可观察事实；评论中的购买、效果、身份或体验主张仍需另行核验。",
        "DataTool兼容表是导出时点的静态查看快照；如需修改数据，请以评论明细和 data 原始文件为准并重新生成。",
        "本文件不包含语义分析、达人画像、商品匹配或商业结论。",
    ])
    style_metadata(intro, 20, [
        ("采集来源", "抖音普通登录页面可见返回"),
        ("隐私模式", comments_manifest.get("privacy_mode") or ""),
        ("开始时间", as_datetime(comments_manifest.get("started_at"))),
        ("完成时间", as_datetime(comments_manifest.get("finished_at"))),
    ])
    set_dimensions(intro, [22, 32, 18, 18, 18, 18])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    book.save(output_path)
    return top_level, replies


def workbook_qa(paths: list[Path]) -> dict[str, Any]:
    results: dict[str, Any] = {}
    error_tokens = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
    for path in paths:
        book = load_workbook(path, data_only=False, read_only=False)
        sheets: dict[str, Any] = {}
        errors: list[str] = []
        id_text_errors: list[str] = []
        for sheet in book.worksheets:
            formulas = 0
            hyperlinks = 0
            id_columns = {
                cell.column
                for cell in sheet[1]
                if isinstance(cell.value, str) and cell.value.strip().upper().endswith("ID")
            }
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas += 1
                    if cell.value in error_tokens:
                        errors.append(f"{sheet.title}!{cell.coordinate}:{cell.value}")
                    if cell.hyperlink is not None:
                        hyperlinks += 1
                    if (
                        cell.row > 1
                        and cell.column in id_columns
                        and cell.value not in (None, "")
                        and (cell.data_type != "s" or cell.number_format != "@")
                    ):
                        id_text_errors.append(
                            f"{sheet.title}!{cell.coordinate}:"
                            f"type={cell.data_type},format={cell.number_format}"
                        )
            sheets[sheet.title] = {
                "rows": sheet.max_row,
                "columns": sheet.max_column,
                "freeze_panes": str(sheet.freeze_panes or ""),
                "tables": sorted(sheet.tables.keys()),
                "formulas": formulas,
                "hyperlinks": hyperlinks,
            }
        results[path.name] = {
            "bytes": path.stat().st_size,
            "sheets": sheets,
            "formula_errors": errors,
            "id_text_errors": id_text_errors,
        }
        book.close()
    return results


def build_explanation(
    output_dir: Path,
    creator: str,
    works_count: int,
    top_level: int,
    replies: int,
    works_manifest: dict[str, Any],
    comments_manifest: dict[str, Any],
) -> None:
    text = (
        "# BrandBAI 抖音基础采集说明\n\n"
        f"- 达人：{creator}\n"
        f"- 作品范围：全部可见置顶作品 + 最近 {int(as_number(works_manifest.get('requested_recent_non_pinned'), 0) or 0)} 条非置顶作品\n"
        f"- 入选作品：{works_count} 条\n"
        f"- 一级评论：{top_level} 条\n"
        f"- 二级回复：{replies} 条\n"
        f"- 作品状态：{display_status(works_manifest.get('status'))}\n"
        f"- 评论状态：{display_status(comments_manifest.get('status'))}\n\n"
        "## 文件说明\n\n"
        "- `01_作品清单.xlsx`：普通阅读版作品与素材清单。\n"
        "- `02_评论明细.xlsx`：普通阅读版评论、DataTool 兼容视图和采集质量。\n"
        "- `03_作品素材`：视频或全部图文、封面与公开原声。\n"
        "- `data`：断点续跑和审计所需原始数据，不作为普通阅读入口。\n\n"
        "## 边界\n\n"
        "“全部”指本次普通登录页面可分页返回且收到终止信号的全部可检索数据，不代表平台内部绝对全量。"
        "页面显示评论数可能同时包含一级评论和其下回复；一级评论中的回复数字段不等于本次实际采集的回复。"
        "本交付不包含语义分析、达人画像、商品匹配或商业结论。\n"
    )
    (output_dir / "04_采集说明.md").write_text(text, encoding="utf-8")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build BrandBAI ordinary-delivery workbooks")
    parser.add_argument("works_json")
    parser.add_argument("works_manifest")
    parser.add_argument("comments_csv")
    parser.add_argument("comments_manifest")
    parser.add_argument("output_dir")
    parser.add_argument("--qa-dir", default="")
    return parser


def main(argv: list[str] | None = None) -> int:
    configure_output()
    args = build_parser().parse_args(argv)
    works_path = Path(args.works_json).expanduser().resolve()
    works_manifest_path = Path(args.works_manifest).expanduser().resolve()
    comments_path = Path(args.comments_csv).expanduser().resolve()
    comments_manifest_path = Path(args.comments_manifest).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    qa_dir = (
        Path(args.qa_dir).expanduser().resolve()
        if args.qa_dir
        else output_dir.parent / f"{output_dir.name}_QA"
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    qa_dir.mkdir(parents=True, exist_ok=True)

    works = load_works(works_path)
    comments = load_comments(comments_path)
    works_manifest = load_json(works_manifest_path)
    comments_manifest = load_json(comments_manifest_path)
    if not isinstance(works_manifest, dict) or not isinstance(comments_manifest, dict):
        raise ValueError("manifest files must contain JSON objects")

    works_output = output_dir / "01_作品清单.xlsx"
    comments_output = output_dir / "02_评论明细.xlsx"
    works_count, asset_record_count, media_file_count = build_works_book(
        works, works_manifest, works_output
    )
    top_level, replies = build_comments_book(works, comments, comments_manifest, comments_output)
    creator = str(works[0].get("author") or "未知达人") if works else "未知达人"
    build_explanation(
        output_dir, creator, works_count, top_level, replies, works_manifest, comments_manifest
    )
    qa = workbook_qa([works_output, comments_output])
    (qa_dir / "workbook_qa.json").write_text(
        json.dumps(qa, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps({
        "creator": creator,
        "works": works_count,
        "assets": media_file_count,
        "assetRecords": asset_record_count,
        "comments": len(comments),
        "topLevel": top_level,
        "replies": replies,
        "outputs": [str(works_output), str(comments_output)],
        "qaDir": str(qa_dir),
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

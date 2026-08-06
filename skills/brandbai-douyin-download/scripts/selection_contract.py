"""Normalize explicit Douyin work selections from JSON or BrandBAI Excel.

The contract is intentionally metadata-only. It never contains cookies, request
headers, browser profile data, or platform signatures. Missing media candidates
are enriched later through the normal visible signed-in work page.
"""

from __future__ import annotations

import json
import re
import urllib.parse
from pathlib import Path
from typing import Any, Iterable


AWEME_ID_RE = re.compile(r"(?:/video/|/note/)(\d{10,})")
SUPPORTED_CONTRACTS = {"brandbai.douyin.selection/v1", ""}


class SelectionContractError(RuntimeError):
    pass


def as_int(value: Any, default: int = 0) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return default


def unique_urls(values: Iterable[Any]) -> list[str]:
    output: list[str] = []
    seen: set[str] = set()
    for value in values:
        url = str(value or "").strip()
        if not url.startswith(("http://", "https://")) or url in seen:
            continue
        seen.add(url)
        output.append(url)
    return output


def normalize_type(value: Any, source_url: str = "") -> str:
    text = str(value or "").strip().lower()
    if text in {"图文", "note", "image", "image_post", "image-post"} or "/note/" in source_url:
        return "图文"
    return "视频"


def source_url_for(aweme_id: str, work_type: str, source_url: Any = "") -> str:
    candidate = str(source_url or "").strip()
    if candidate.startswith(("http://", "https://")) and ("/video/" in candidate or "/note/" in candidate):
        return candidate
    route = "note" if work_type == "图文" else "video"
    return f"https://www.douyin.com/{route}/{aweme_id}"


def work_id_from_url(value: str) -> str:
    match = AWEME_ID_RE.search(value)
    if match:
        return match.group(1)
    query = urllib.parse.parse_qs(urllib.parse.urlparse(value).query)
    modal_id = str((query.get("modal_id") or [""])[0]).strip()
    return modal_id if modal_id.isdigit() and len(modal_id) >= 10 else ""


def nested_urls(value: Any) -> list[str]:
    if isinstance(value, str):
        return unique_urls([value])
    if isinstance(value, list):
        return unique_urls(value)
    if not isinstance(value, dict):
        return []
    collected: list[Any] = []
    for key in ("urls", "url_list", "urlList", "candidates"):
        item = value.get(key)
        if isinstance(item, list):
            collected.extend(item)
    for key in ("url", "download_url", "downloadUrl"):
        item = value.get(key)
        if isinstance(item, str):
            collected.append(item)
    return unique_urls(collected)


def normalize_row(row: dict[str, Any], rank: int, defaults: dict[str, Any] | None = None) -> dict[str, Any]:
    defaults = defaults or {}
    source_url = str(
        row.get("source_url") or row.get("作品链接") or row.get("url") or ""
    ).strip()
    url_aweme_id = work_id_from_url(source_url)
    aweme_id = str(
        row.get("aweme_id") or row.get("awemeId") or row.get("作品ID") or url_aweme_id
    ).strip()
    if not aweme_id or not aweme_id.isdigit():
        raise SelectionContractError(f"Selection row {rank} is missing a valid Douyin work ID")
    work_type = normalize_type(row.get("type") or row.get("类型"), source_url)
    media = row.get("media") if isinstance(row.get("media"), dict) else {}
    statistics = row.get("statistics") if isinstance(row.get("statistics"), dict) else {}
    images = media.get("images") or row.get("image_candidates") or []
    image_candidates: list[list[str]] = []
    if isinstance(images, list):
        for image in images:
            urls = nested_urls(image)
            if urls:
                image_candidates.append(urls)
    return {
        "aweme_id": aweme_id,
        "type": work_type,
        "author": str(row.get("author") or row.get("作者") or "").strip(),
        "title": str(row.get("title") or row.get("标题/发布文案") or row.get("发布文案") or "").strip(),
        "create_time": as_int(row.get("create_time") or row.get("createTime")),
        "publish_time": str(row.get("publish_time") or row.get("发布时间") or "").strip(),
        "digg_count": as_int(row.get("digg_count") or row.get("点赞数") or statistics.get("digg_count")),
        "share_count": as_int(row.get("share_count") or row.get("分享数") or statistics.get("share_count")),
        "comment_count": as_int(row.get("comment_count") or row.get("评论数") or statistics.get("comment_count")),
        "collect_count": as_int(row.get("collect_count") or row.get("收藏数") or statistics.get("collect_count")),
        "recommend_count": row.get("recommend_count", row.get("推荐数", statistics.get("recommend_count"))),
        "is_pinned": str(row.get("is_pinned") or row.get("是否置顶") or "").strip().lower() in {"1", "true", "yes", "是"},
        "source_url": source_url_for(aweme_id, work_type, source_url),
        "source_page_type": str(row.get("source_page_type") or row.get("来源页面类型") or defaults.get("page_type") or "selection").strip(),
        "source_keyword": str(row.get("source_keyword") or row.get("来源关键词") or defaults.get("keyword") or "").strip(),
        "source_rank": as_int(row.get("source_rank") or row.get("来源排序"), rank),
        "selection_reason": str(row.get("selection_reason") or defaults.get("selection_reason") or "手动选择").strip(),
        "selection_rank": rank,
        "_video_urls": nested_urls(media.get("video") or row.get("video_candidates")),
        "_cover_urls": nested_urls(media.get("cover") or row.get("cover_candidates")),
        "_music_urls": nested_urls(media.get("audio") or media.get("music") or row.get("music_candidates")),
        "_image_urls": image_candidates,
    }


def deduplicate(rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows:
        aweme_id = str(row.get("aweme_id") or "")
        if not aweme_id or aweme_id in seen:
            continue
        seen.add(aweme_id)
        row["selection_rank"] = len(output) + 1
        output.append(row)
    return output


def load_json(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    try:
        payload: Any = json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError) as exc:
        raise SelectionContractError(f"Cannot read selection JSON: {path}") from exc
    metadata: dict[str, Any] = {}
    if isinstance(payload, dict):
        contract = str(payload.get("contract") or "").strip()
        if contract not in SUPPORTED_CONTRACTS:
            raise SelectionContractError(f"Unsupported selection contract: {contract}")
        source = payload.get("source") if isinstance(payload.get("source"), dict) else {}
        selection = payload.get("selection") if isinstance(payload.get("selection"), dict) else {}
        metadata = {
            "contract": contract or "brandbai.douyin.selection/v1",
            "page_type": source.get("page_type") or "selection",
            "page_url": source.get("page_url") or "",
            "keyword": source.get("keyword") or "",
            "captured_at": source.get("captured_at") or "",
            "selection_mode": selection.get("mode") or "manual",
            "selection_reason": selection.get("label") or "手动选择",
            "download": payload.get("download") if isinstance(payload.get("download"), dict) else {},
        }
        rows = payload.get("works")
    else:
        rows = payload
    if not isinstance(rows, list):
        raise SelectionContractError("Selection JSON must contain a works list")
    normalized = [normalize_row(row, index, metadata) for index, row in enumerate(rows, 1) if isinstance(row, dict)]
    normalized = deduplicate(normalized)
    if not normalized:
        raise SelectionContractError(f"No usable works found in selection JSON: {path}")
    return normalized, metadata


def load_xlsx(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SelectionContractError("openpyxl is required to read selection Excel files") from exc
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surface workbook parsing failures
        raise SelectionContractError(f"Cannot read selection Excel: {path}") from exc
    try:
        if "作品清单" not in workbook.sheetnames:
            raise SelectionContractError("Selection Excel must contain a 作品清单 sheet")
        sheet = workbook["作品清单"]
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(iterator, [])]
        if "作品ID" not in headers:
            raise SelectionContractError("作品清单 is missing the 作品ID column")
        rows: list[dict[str, Any]] = []
        for values in iterator:
            row = {headers[index]: values[index] for index in range(min(len(headers), len(values))) if headers[index]}
            if str(row.get("作品ID") or "").strip():
                rows.append(row)
        normalized = deduplicate(normalize_row(row, index) for index, row in enumerate(rows, 1))
        if not normalized:
            raise SelectionContractError(f"No usable works found in selection Excel: {path}")
        page_types = {row.get("source_page_type") for row in normalized if row.get("source_page_type")}
        keywords = {row.get("source_keyword") for row in normalized if row.get("source_keyword")}
        return normalized, {
            "contract": "brandbai.douyin.selection/v1",
            "page_type": next(iter(page_types)) if len(page_types) == 1 else "mixed",
            "page_url": "",
            "keyword": next(iter(keywords)) if len(keywords) == 1 else "",
            "captured_at": "",
            "selection_mode": "manual_excel",
            "selection_reason": "插件作品清单",
            "download": {},
        }
    finally:
        workbook.close()


def load_selection(path_value: str | Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    path = Path(path_value).expanduser().resolve()
    if not path.is_file():
        raise SelectionContractError(f"Selection file not found: {path}")
    if path.suffix.lower() == ".json":
        return load_json(path)
    if path.suffix.lower() == ".xlsx":
        return load_xlsx(path)
    raise SelectionContractError("--selection-file must be .json or .xlsx")


def seed_from_url(url: str, rank: int) -> dict[str, Any]:
    value = str(url or "").strip()
    aweme_id = work_id_from_url(value)
    if not aweme_id:
        raise SelectionContractError(f"Explicit work URL is missing /video/, /note/, or modal_id: {value}")
    work_type = "图文" if "/note/" in value else "视频"
    return normalize_row(
        {
            "aweme_id": aweme_id,
            "type": work_type,
            "source_url": value,
            "selection_reason": "明确作品",
        },
        rank,
    )
